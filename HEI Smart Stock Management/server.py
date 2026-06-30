"""
MECP Part Order Management System — Backend Server
FastAPI + WebSocket + Uvicorn + SQLite
Local Network — โรงงาน MECP
"""

import json
import os
import re
import sqlite3
from datetime import datetime
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Query, Request, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, FileResponse
import uvicorn

app = FastAPI(title="MECP Part Order System")
app.mount("/static", StaticFiles(directory="public"), name="static")

# ───────────────────────────────────────────────
# JSON Store (memory + disk)
# ───────────────────────────────────────────────
DATA_FILE = "orders.json"
orders: list = []
order_counter: int = 1

ORDERS_LIMIT = 500

def load_data():
    global orders, order_counter
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                saved = json.load(f)
                all_orders = saved.get("orders", [])
                order_counter = saved.get("counter", 1)
            # เก็บเฉพาะ 500 รายการล่าสุดใน memory — ที่เหลืออยู่ใน SQLite
            orders = sorted(all_orders, key=lambda o: o.get("timestamp", ""), reverse=True)[:ORDERS_LIMIT]
            print(f"[JSON] โหลด {len(orders)} Orders (ทั้งหมด {len(all_orders)} ใน file)")
        except Exception as e:
            print(f"[JSON] Error: {e}")

def save_json():
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump({"orders": orders, "counter": order_counter}, f, ensure_ascii=False, indent=2)

load_data()

# ───────────────────────────────────────────────
# SQLite Database
# ───────────────────────────────────────────────
DB_FILE = "mecp.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "CREATE TABLE IF NOT EXISTS orders ("
        "id TEXT PRIMARY KEY, part TEXT, model TEXT, qty INTEGER, "
        "priority TEXT, remark TEXT, date TEXT, time TEXT, "
        "timestamp TEXT, status TEXT, confirm_time TEXT, work_status TEXT, "
        "lot TEXT, line_num TEXT, seq TEXT, pro_month TEXT, "
        "employee TEXT, item_code TEXT, raw_qr TEXT, "
        "beg_qty INTEGER, teem_qty INTEGER)"
    )
    conn.commit()
    conn.close()
    print(f"[DB] SQLite ready: {DB_FILE}")

def db_migrate_columns():
    new_cols = [
        ("lot", "TEXT"), ("line_num", "TEXT"), ("seq", "TEXT"),
        ("pro_month", "TEXT"), ("employee", "TEXT"), ("item_code", "TEXT"),
        ("raw_qr", "TEXT"), ("beg_qty", "INTEGER"), ("teem_qty", "INTEGER"),
        ("confirm_by", "TEXT"), ("confirm_date", "TEXT"), ("need_by", "TEXT"), ("machine", "TEXT"),
        ("floor_remark", "TEXT"),
        ("receive_time", "TEXT"), ("receive_date", "TEXT"), ("receive_by", "TEXT"),
        ("qty_adjusted", "INTEGER"), ("original_qty", "INTEGER"),
        ("ubend_sub", "TEXT"),
        ("item_cond", "TEXT"),
        ("prod_order", "TEXT"),
    ]
    conn = sqlite3.connect(DB_FILE)
    existing = {row[1] for row in conn.execute("PRAGMA table_info(orders)")}
    for col, typ in new_cols:
        if col not in existing:
            conn.execute(f"ALTER TABLE orders ADD COLUMN {col} {typ}")
            print(f"[DB] Added column: {col}")
    conn.commit()
    conn.close()

def db_insert(o: dict):
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "INSERT OR IGNORE INTO orders "
            "(id,part,model,qty,priority,remark,date,time,timestamp,status,"
            "confirm_time,work_status,lot,line_num,seq,pro_month,employee,"
            "item_code,item_cond,raw_qr,beg_qty,teem_qty,need_by,machine,qty_adjusted,original_qty,ubend_sub,prod_order) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (o["id"], o["part"], o["model"], int(o["qty"]),
             o.get("priority","normal"), o.get("remark",""),
             o.get("date",""), o.get("time",""), o.get("timestamp",""),
             o.get("status","pending"), o.get("confirmTime"), o.get("workStatus"),
             o.get("lot"), o.get("lineNum"), o.get("seq"), o.get("proMonth"),
             o.get("employee"), o.get("itemCode"), o.get("itemCond"), o.get("rawQr"),
             o.get("begQty"), o.get("teemQty"), o.get("needBy"), o.get("machine"),
             1 if o.get("qtyAdjusted") else 0, o.get("originalQty"), o.get("ubendSub"),
             o.get("prodOrder"))
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Insert error: {e}")

def db_confirm(order_id: str, work_status: str, confirm_time: str, confirm_by: str = None, confirm_date: str = None, floor_remark: str = None):
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "UPDATE orders SET status='confirmed', work_status=?, confirm_time=?, confirm_by=?, confirm_date=?, floor_remark=? WHERE id=?",
            (work_status, confirm_time, confirm_by, confirm_date, floor_remark, order_id)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Confirm error: {e}")

def db_receive(order_id: str, receive_time: str, receive_date: str, receive_by: str):
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "UPDATE orders SET work_status='รับแล้ว', receive_time=?, receive_date=?, receive_by=? WHERE id=?",
            (receive_time, receive_date, receive_by, order_id)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Receive error: {e}")

def db_unconfirm(order_id: str):
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "UPDATE orders SET status='pending', work_status=NULL, confirm_time=NULL, floor_remark=NULL WHERE id=?",
            (order_id,)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Unconfirm error: {e}")

def db_query(status=None, part=None, priority=None, search=None, limit=500):
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    q = "SELECT * FROM orders WHERE 1=1"
    p = []
    if status and status != "all":
        q += " AND status=?"; p.append(status)
    if part and part != "all":
        q += " AND part=?"; p.append(part)
    if priority and priority != "all":
        q += " AND priority=?"; p.append(priority)
    if search:
        q += " AND (id LIKE ? OR model LIKE ? OR part LIKE ?)"; p.extend([f"%{search}%"]*3)
    q += " ORDER BY timestamp DESC LIMIT ?"; p.append(limit)
    rows = conn.execute(q, p).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_order_stock_qty(order: dict) -> int:
    """จำนวนที่ใช้หัก stock: ถ้ามีการปรับจำนวน ให้ใช้ qty ที่ปรับแล้วก่อน."""
    try:
        is_adjusted = bool(order.get("qtyAdjusted") or order.get("qty_adjusted"))
        if is_adjusted:
            return int(order.get("qty") or 0)
        if order.get("part") == "U-Bend" and (order.get("teemQty") or order.get("teem_qty")):
            return int(order.get("teemQty") or order.get("teem_qty") or 0)
        return int(order.get("qty") or 0)
    except (TypeError, ValueError):
        return 0

def db_get_order(order_id: str):
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def _th_date_to_iso(d: str) -> str:
    try:
        p = str(d or "").strip().split("/")
        return f"{p[2]}-{p[1].zfill(2)}-{p[0].zfill(2)}"
    except Exception:
        return ""

def db_orders_for_delete(before: str = None):
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    rows = [dict(r) for r in conn.execute("SELECT * FROM orders").fetchall()]
    conn.close()
    if not before:
        return rows
    return [r for r in rows if _th_date_to_iso(r.get("date", "")) < before]

def db_stats():
    conn = sqlite3.connect(DB_FILE)
    s = {}
    s["total"]      = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
    s["pending"]    = conn.execute("SELECT COUNT(*) FROM orders WHERE status='pending'").fetchone()[0]
    s["confirmed"]  = conn.execute("SELECT COUNT(*) FROM orders WHERE status='confirmed'").fetchone()[0]
    s["urgent"]     = conn.execute("SELECT COUNT(*) FROM orders WHERE priority='urgent'").fetchone()[0]
    s["side_plate"] = conn.execute("SELECT COUNT(*) FROM orders WHERE part='Side Plate'").fetchone()[0]
    s["ubend"]      = conn.execute("SELECT COUNT(*) FROM orders WHERE part='U-Bend'").fetchone()[0]
    s["total_qty"]  = conn.execute("SELECT COALESCE(SUM(qty),0) FROM orders").fetchone()[0]
    conn.close()
    return s

def db_migrate():
    if not orders:
        return
    n = 0
    for o in orders:
        try:
            db_insert(o)
            n += 1
        except Exception:
            pass
    if n:
        print(f"[DB] Migrated {n} orders from JSON to SQLite")

init_db()
db_migrate_columns()
db_migrate()

# ───────────────────────────────────────────────
# INCOMING DATABASE  (หน้ารับ Part — แยกจาก mecp.db)
# ───────────────────────────────────────────────
INCOMING_DB = "incoming.db"

def incoming_init_db():
    conn = sqlite3.connect(INCOMING_DB)
    conn.execute(
        "CREATE TABLE IF NOT EXISTS incoming ("
        "id TEXT PRIMARY KEY,"
        "do_num TEXT,"
        "part_no TEXT,"
        "item_code TEXT,"
        "qty INTEGER,"
        "orig_qty INTEGER,"
        "unit TEXT DEFAULT 'PC',"
        "line_num TEXT,"
        "seq TEXT,"
        "location TEXT,"
        "due_date TEXT,"
        "due_time TEXT,"
        "pro_month TEXT,"
        "raw_qr TEXT,"
        "employee TEXT,"
        "receive_date TEXT,"
        "receive_time TEXT,"
        "timestamp TEXT"
        ")"
    )
    # ตารางหักถาวร: เก็บจำนวนที่ถูกลบแล้วแต่ไม่คืน stock
    conn.execute(
        "CREATE TABLE IF NOT EXISTS deductions ("
        "id TEXT PRIMARY KEY,"
        "part_no TEXT NOT NULL,"
        "qty INTEGER NOT NULL,"
        "reason TEXT,"
        "ref_order_id TEXT,"
        "date TEXT,"
        "timestamp TEXT"
        ")"
    )
    # ตาราง lookup: item_code → description สกัดจาก raw_qr
    conn.execute(
        "CREATE TABLE IF NOT EXISTS item_descriptions ("
        "item_code TEXT PRIMARY KEY,"
        "description TEXT,"
        "part_type TEXT,"
        "updated_at TEXT"
        ")"
    )
    # ตาราง vendors: vendor_code → vendor_name
    conn.execute(
        "CREATE TABLE IF NOT EXISTS vendors ("
        "vendor_code TEXT PRIMARY KEY,"
        "vendor_name TEXT NOT NULL,"
        "updated_at TEXT"
        ")"
    )
    conn.commit()
    # migration: เพิ่มคอลัมน์ orig_qty ให้ DB เดิม + เติมค่าจาก qty ปัจจุบันถ้ายังว่าง
    cols = [r[1] for r in conn.execute("PRAGMA table_info(incoming)").fetchall()]
    if "orig_qty" not in cols:
        conn.execute("ALTER TABLE incoming ADD COLUMN orig_qty INTEGER")
    conn.execute("UPDATE incoming SET orig_qty=qty WHERE orig_qty IS NULL")
    conn.commit()
    conn.close()
    print(f"[INCOMING] SQLite ready: {INCOMING_DB}")

_DEFAULT_VENDORS = {
    "100840": "KISHIMOTO",
    "100471": "S.Y.K.",
    "100324": "TAP / Thamaruk Auto Part",
    "100303": "TSE",
    "100393": "IPC",
    "100403": "IL JIN",
    "100327": "KAISE",
    "100311": "AMAGA",
}

def seed_vendors():
    conn = sqlite3.connect(INCOMING_DB)
    count = conn.execute("SELECT COUNT(*) FROM vendors").fetchone()[0]
    if count == 0:
        now = datetime.now().isoformat()
        for code, name in _DEFAULT_VENDORS.items():
            conn.execute(
                "INSERT OR IGNORE INTO vendors (vendor_code, vendor_name, updated_at) VALUES (?,?,?)",
                (code, name, now)
            )
        conn.commit()
        print(f"[VENDORS] Seeded {len(_DEFAULT_VENDORS)} default vendors")
    conn.close()


import re as _re

def _extract_desc_from_qr(raw_qr: str, item_code: str):
    """ดึง description จาก raw_qr: {item_code}  {DESC}  {MMYYYY}"""
    try:
        m = _re.search(_re.escape(item_code) + r'\s{2,}(.+?)\s{2,}\d{6}', raw_qr)
        return m.group(1).strip() if m else None
    except Exception:
        return None

def _upsert_item_desc(item_code: str, description, part_type: str):
    try:
        conn = sqlite3.connect(INCOMING_DB)
        conn.execute(
            "INSERT INTO item_descriptions (item_code, description, part_type, updated_at) "
            "VALUES (?,?,?,?) ON CONFLICT(item_code) DO UPDATE SET "
            "description=excluded.description, part_type=excluded.part_type, updated_at=excluded.updated_at",
            (item_code, description, part_type, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

def build_item_descriptions():
    """สร้าง item_descriptions จาก raw_qr ทั้งหมดใน mecp.db"""
    try:
        mecp = sqlite3.connect(DB_FILE)
        mecp.row_factory = sqlite3.Row
        rows = mecp.execute(
            "SELECT item_code, part, raw_qr FROM orders "
            "WHERE raw_qr IS NOT NULL AND item_code IS NOT NULL AND item_code != '' "
            "GROUP BY item_code"
        ).fetchall()
        mecp.close()
        inc = sqlite3.connect(INCOMING_DB)
        now = datetime.now().isoformat()
        count = 0
        for r in rows:
            ic   = str(r["item_code"] or "").strip()
            desc = _extract_desc_from_qr(r["raw_qr"] or "", ic)
            if not ic: continue
            inc.execute(
                "INSERT INTO item_descriptions (item_code, description, part_type, updated_at) "
                "VALUES (?,?,?,?) ON CONFLICT(item_code) DO UPDATE SET "
                "description=excluded.description, part_type=excluded.part_type, updated_at=excluded.updated_at",
                (ic, desc, r["part"], now)
            )
            count += 1
        inc.commit()
        inc.close()
        print(f"[ITEM-DESC] Built {count} entries")
    except Exception as e:
        print(f"[ITEM-DESC] Error: {e}")

def create_deduction(order: dict):
    """บันทึกรายการหักถาวรเมื่อลบ order แบบ 'คง stock'"""
    import uuid
    part_no = str(order.get("item_code") or order.get("itemCode") or "").strip()
    qty = get_order_stock_qty(order)
    if not part_no or qty <= 0:
        return
    now = datetime.now()
    conn = sqlite3.connect(INCOMING_DB)
    conn.execute(
        "INSERT OR IGNORE INTO deductions (id, part_no, qty, reason, ref_order_id, date, timestamp) "
        "VALUES (?,?,?,?,?,?,?)",
        (str(uuid.uuid4()), part_no, qty, "order_deleted",
         order.get("id",""), now.strftime("%d/%m/%Y"), now.isoformat())
    )
    conn.commit()
    conn.close()
    print(f"[DEDUCTION] {part_no} qty={qty} ref={order.get('id','')}")

def incoming_insert(rec: dict):
    conn = sqlite3.connect(INCOMING_DB)
    conn.execute(
        "INSERT OR IGNORE INTO incoming "
        "(id,do_num,part_no,item_code,qty,orig_qty,unit,line_num,seq,location,"
        " due_date,due_time,pro_month,raw_qr,employee,receive_date,receive_time,timestamp) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            rec["id"],
            rec.get("doNum",""),
            rec.get("partNo",""),
            rec.get("itemCode",""),
            int(rec.get("qty") or 0),
            int(rec.get("qty") or 0),
            rec.get("unit","PC"),
            rec.get("lineNum",""),
            rec.get("seq",""),
            rec.get("location",""),
            rec.get("dueDate",""),
            rec.get("dueTime",""),
            rec.get("proMonth",""),
            rec.get("rawQr",""),
            rec.get("employee",""),
            rec.get("receiveDate",""),
            rec.get("receiveTime",""),
            rec.get("timestamp",""),
        )
    )
    conn.commit()
    conn.close()

def incoming_query(line=None, search=None, date_from=None, date_to=None, limit=500):
    conn = sqlite3.connect(INCOMING_DB)
    conn.row_factory = sqlite3.Row
    q = ("SELECT i.*, COALESCE(d.description, '') AS description "
         "FROM incoming i "
         "LEFT JOIN item_descriptions d "
         "ON COALESCE(NULLIF(i.part_no,''), i.item_code) = d.item_code "
         "WHERE 1=1")
    p = []
    if line and line != "all":
        q += " AND i.line_num=?"; p.append(line)
    if search:
        # i.* ถูก JOIN กับ item_descriptions ที่มี item_code เหมือนกัน — ต้องระบุ alias i. กันชื่อชนกัน
        q += (" AND (i.part_no LIKE ? OR i.item_code LIKE ? OR i.do_num LIKE ? "
              "OR i.seq LIKE ? OR i.location LIKE ? OR i.employee LIKE ?)")
        p.extend([f"%{search}%"]*6)
    if date_from:
        q += " AND i.receive_date>=?"; p.append(date_from)
    if date_to:
        q += " AND i.receive_date<=?"; p.append(date_to)
    q += " ORDER BY timestamp DESC LIMIT ?"; p.append(limit)
    rows = conn.execute(q, p).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def incoming_stats():
    conn = sqlite3.connect(INCOMING_DB)
    s = {}
    s["total"]      = conn.execute("SELECT COUNT(*) FROM incoming").fetchone()[0]
    s["total_qty"]  = conn.execute("SELECT COALESCE(SUM(qty),0) FROM incoming").fetchone()[0]
    s["today"]      = conn.execute(
        "SELECT COUNT(*) FROM incoming WHERE receive_date=?",
        (datetime.now().strftime("%d/%m/%Y"),)
    ).fetchone()[0]
    for ln in ["A","B","G","T","T2","H"]:
        s[f"line_{ln}"] = conn.execute(
            "SELECT COUNT(*) FROM incoming WHERE line_num=?", (ln,)
        ).fetchone()[0]
    conn.close()
    return s

incoming_init_db()
seed_vendors()
build_item_descriptions()

# ───────────────────────────────────────────────
# REST API — Incoming endpoints
# ───────────────────────────────────────────────
@app.post("/api/incoming")
async def api_incoming_create(request: Request):
    try:
        body = await request.json()
        now = datetime.now()
        rec_id = f"INC-{now.strftime('%Y%m%d-%H%M%S')}-{str(int(now.timestamp()*1000))[-4:]}"
        body["id"]          = rec_id
        body["receiveDate"] = now.strftime("%d/%m/%Y")
        body["receiveTime"] = now.strftime("%H:%M:%S")
        body["timestamp"]   = now.isoformat()
        incoming_insert(body)
        return JSONResponse({"ok": True, "id": rec_id, "record": body})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# หัวตาราง Excel (normalize แล้ว) → ชื่อ field ภายใน — รองรับทั้งไทยและอังกฤษ
# normalize = ตัวพิมพ์เล็ก + ตัดช่องว่าง/ขีด/ขีดล่าง/สแลชออก ให้ "Part No" = "part_no" = "partno"
_EXCEL_HEADER_MAP = {
    "partno": "partNo", "part": "partNo", "partnumber": "partNo",
    "itemcode": "itemCode", "item": "itemCode", "vendorcode": "itemCode", "vendor": "itemCode",
    "จำนวน": "qty", "qty": "qty", "quantity": "qty", "amount": "qty",
    "unit": "unit", "หน่วย": "unit",
    "line": "lineNum", "linenum": "lineNum", "ไลน์": "lineNum",
    "seq": "seq", "sequence": "seq",
    "location": "location", "loc": "location", "ตำแหน่ง": "location",
    "duedate": "dueDate", "duetime": "dueTime",
    "donumber": "doNum", "do": "doNum", "donum": "doNum", "dono": "doNum",
    "promonth": "proMonth",
    "พนักงาน": "employee", "employee": "employee", "emp": "employee", "staff": "employee",
    "วันที่รับ": "receiveDate", "receivedate": "receiveDate", "date": "receiveDate",
    "เวลา": "receiveTime", "receivetime": "receiveTime", "time": "receiveTime",
    "rawqr": "rawQr", "qr": "rawQr",
    "recordid": "id", "id": "id",
}

def _norm_header(h):
    s = str(h or "").strip().lower()
    for ch in (" ", "_", "-", "/", ".", "#"):
        s = s.replace(ch, "")
    return s

@app.post("/api/incoming/import")
async def api_incoming_import(file: UploadFile = File(...)):
    """นำเข้าข้อมูล incoming จากไฟล์ Excel (.xlsx) — แมปคอลัมน์จากหัวตาราง รองรับรูปแบบเดียวกับ Export CSV"""
    try:
        import openpyxl, io, uuid
        data = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        header = next(rows, None)
        if not header:
            return JSONResponse({"ok": False, "error": "ไฟล์ว่างเปล่า"}, status_code=400)
        # คอลัมน์ index → field
        col_map = {}
        for idx, h in enumerate(header):
            field = _EXCEL_HEADER_MAP.get(_norm_header(h))
            if field:
                col_map[idx] = field
        if "partNo" not in col_map.values() and "itemCode" not in col_map.values():
            return JSONResponse(
                {"ok": False, "error": "ไม่พบคอลัมน์ Part No หรือ Item Code ในไฟล์"},
                status_code=400)

        now = datetime.now()
        imported = skipped = 0
        for r in rows:
            rec = {}
            for idx, field in col_map.items():
                if idx < len(r):
                    rec[field] = r[idx]
            part = str(rec.get("partNo") or rec.get("itemCode") or "").strip()
            if not part:                       # ข้ามแถวว่าง
                skipped += 1
                continue
            # qty → int (รองรับค่าทศนิยม/มี comma)
            try:
                rec["qty"] = int(float(str(rec.get("qty") or 0).replace(",", "")))
            except (ValueError, TypeError):
                rec["qty"] = 0
            # ทำความสะอาดค่า: แปลงทุกอย่างเป็น str ตัดช่องว่าง
            for k in list(rec.keys()):
                if k != "qty" and rec[k] is not None:
                    rec[k] = str(rec[k]).strip()
            rec.setdefault("unit", "PC")
            if not rec.get("id"):
                rec["id"] = f"INC-{now.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"
            rec.setdefault("receiveDate", now.strftime("%d/%m/%Y"))
            rec.setdefault("receiveTime", now.strftime("%H:%M:%S"))
            rec["timestamp"] = now.isoformat()
            incoming_insert(rec)
            imported += 1
        wb.close()
        return JSONResponse({"ok": True, "imported": imported, "skipped": skipped})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

UBEND_USAGE_DB = "ubend_usage.db"

@app.get("/api/ubend-usage/{item_code}")
async def api_ubend_usage(item_code: str):
    try:
        conn = sqlite3.connect(UBEND_USAGE_DB)
        rows = conn.execute(
            """SELECT u.ubend_code, u.description, u.unit, iu.qty AS usage_qty
               FROM item_ubend_usage iu
               JOIN ubend_types u ON iu.ubend_code = u.ubend_code
               WHERE iu.item_code = ?
               ORDER BY iu.qty DESC""",
            (item_code,)
        ).fetchall()
        conn.close()
        if not rows:
            return {"found": False, "items": []}
        return {"found": True, "items": [
            {"ubend_code": r[0], "description": r[1], "unit": r[2], "usage_qty": float(r[3])}
            for r in rows
        ]}
    except Exception as e:
        return {"found": False, "items": [], "error": str(e)}

# ─────────────────────────────────────────────────────────────────
#  U-Bend Usage DB — CRUD endpoints
# ─────────────────────────────────────────────────────────────────

def ubend_conn():
    c = sqlite3.connect(UBEND_USAGE_DB)
    c.row_factory = sqlite3.Row
    return c

@app.get("/api/ubend-db/stats")
async def ubend_stats():
    try:
        c = ubend_conn()
        stats = {
            "items":      c.execute("SELECT COUNT(*) FROM items").fetchone()[0],
            "ubend_types": c.execute("SELECT COUNT(*) FROM ubend_types").fetchone()[0],
            "usage":      c.execute("SELECT COUNT(*) FROM item_ubend_usage").fetchone()[0],
        }
        c.close()
        return JSONResponse({"ok": True, **stats})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ── item_ubend_usage ──────────────────────────────────────────────
@app.get("/api/ubend-db/usage")
async def ubend_usage_list(search: str = Query(""), limit: int = Query(100), offset: int = Query(0)):
    try:
        c = ubend_conn()
        base = """FROM item_ubend_usage iu
                  LEFT JOIN items i ON iu.item_code = i.item_code
                  LEFT JOIN ubend_types u ON iu.ubend_code = u.ubend_code"""
        where, p = "", []
        if search:
            where = " WHERE (iu.item_code LIKE ? OR iu.ubend_code LIKE ? OR i.model LIKE ? OR u.description LIKE ?)"
            p = [f"%{search}%"] * 4
        total = c.execute(f"SELECT COUNT(*) {base}{where}", p).fetchone()[0]
        rows  = c.execute(
            f"""SELECT iu.usage_id, iu.item_code, i.model AS item_model,
                       iu.ubend_code, u.description, u.unit, iu.qty
                {base}{where}
                ORDER BY iu.item_code, iu.qty DESC
                LIMIT ? OFFSET ?""",
            p + [limit, offset]
        ).fetchall()
        c.close()
        return JSONResponse({"ok": True, "total": total,
                             "rows": [dict(r) for r in rows]})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/ubend-db/usage")
async def ubend_usage_add(req: Request):
    try:
        d = await req.json()
        item_code  = str(d.get("item_code","")).strip().upper()
        ubend_code = str(d.get("ubend_code","")).strip().upper()
        qty        = float(d.get("qty", 1))
        if not item_code or not ubend_code:
            return JSONResponse({"ok": False, "error": "item_code และ ubend_code ต้องระบุ"}, status_code=400)
        c = ubend_conn()
        # check duplicate
        exists = c.execute("SELECT 1 FROM item_ubend_usage WHERE item_code=? AND ubend_code=?",
                           (item_code, ubend_code)).fetchone()
        if exists:
            c.close()
            return JSONResponse({"ok": False, "error": "รายการนี้มีอยู่แล้ว"}, status_code=409)
        c.execute("INSERT INTO item_ubend_usage (item_code,ubend_code,qty,source_row,source_col_letter) VALUES (?,?,?,0,'M')",
                  (item_code, ubend_code, qty))
        usage_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        c.commit(); c.close()
        return JSONResponse({"ok": True, "usage_id": usage_id})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.put("/api/ubend-db/usage/{usage_id}")
async def ubend_usage_update(usage_id: int, req: Request):
    try:
        d    = await req.json()
        qty  = float(d.get("qty", 1))
        c    = ubend_conn()
        c.execute("UPDATE item_ubend_usage SET qty=? WHERE usage_id=?", (qty, usage_id))
        c.commit(); c.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.delete("/api/ubend-db/usage/{usage_id}")
async def ubend_usage_delete(usage_id: int):
    try:
        c = ubend_conn()
        c.execute("DELETE FROM item_ubend_usage WHERE usage_id=?", (usage_id,))
        c.commit(); c.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ── items (Item COND) ────────────────────────────────────────────
@app.get("/api/ubend-db/items")
async def ubend_items_list(search: str = Query("")):
    try:
        c = ubend_conn()
        q, p = "SELECT item_code, model, status FROM items", []
        if search:
            q += " WHERE item_code LIKE ? OR model LIKE ?"; p = [f"%{search}%"]*2
        q += " ORDER BY item_code"
        rows = c.execute(q, p).fetchall()
        c.close()
        return JSONResponse({"ok": True, "rows": [dict(r) for r in rows]})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/ubend-db/items")
async def ubend_items_add(req: Request):
    try:
        d          = await req.json()
        item_code  = str(d.get("item_code","")).strip().upper()
        model      = str(d.get("model","")).strip()
        status     = str(d.get("status","o")).strip()
        if not item_code:
            return JSONResponse({"ok": False, "error": "ต้องระบุ Item COND"}, status_code=400)
        c = ubend_conn()
        exists = c.execute("SELECT 1 FROM items WHERE item_code=?", (item_code,)).fetchone()
        if exists:
            c.close()
            return JSONResponse({"ok": False, "error": "Item COND นี้มีอยู่แล้ว"}, status_code=409)
        c.execute("INSERT INTO items (item_code,model,status,source_row) VALUES (?,?,?,0)",
                  (item_code, model, status))
        c.commit(); c.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.delete("/api/ubend-db/items/{item_code}")
async def ubend_items_delete(item_code: str):
    try:
        c = ubend_conn()
        c.execute("DELETE FROM items WHERE item_code=?", (item_code,))
        c.commit(); c.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ── ubend_types (Item Ubend) ─────────────────────────────────────
@app.get("/api/ubend-db/ubend-types")
async def ubend_types_list(search: str = Query("")):
    try:
        c = ubend_conn()
        q, p = "SELECT ubend_code, description, unit FROM ubend_types", []
        if search:
            q += " WHERE ubend_code LIKE ? OR description LIKE ?"; p = [f"%{search}%"]*2
        q += " ORDER BY ubend_code"
        rows = c.execute(q, p).fetchall()
        c.close()
        return JSONResponse({"ok": True, "rows": [dict(r) for r in rows]})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/ubend-db/ubend-types")
async def ubend_types_add(req: Request):
    try:
        d           = await req.json()
        ubend_code  = str(d.get("ubend_code","")).strip().upper()
        description = str(d.get("description","")).strip()
        unit        = str(d.get("unit","(PCS.)")).strip()
        if not ubend_code:
            return JSONResponse({"ok": False, "error": "ต้องระบุ Item Ubend Code"}, status_code=400)
        c = ubend_conn()
        exists = c.execute("SELECT 1 FROM ubend_types WHERE ubend_code=?", (ubend_code,)).fetchone()
        if exists:
            c.close()
            return JSONResponse({"ok": False, "error": "Ubend Code นี้มีอยู่แล้ว"}, status_code=409)
        max_idx = c.execute("SELECT COALESCE(MAX(source_col_index),0)+1 FROM ubend_types").fetchone()[0]
        c.execute("INSERT INTO ubend_types (ubend_code,description,unit,source_col_letter,source_col_index) VALUES (?,?,?,'M',?)",
                  (ubend_code, description, unit, max_idx))
        c.commit(); c.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.delete("/api/ubend-db/ubend-types/{ubend_code}")
async def ubend_types_delete(ubend_code: str):
    try:
        c = ubend_conn()
        c.execute("DELETE FROM ubend_types WHERE ubend_code=?", (ubend_code,))
        c.commit(); c.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/incoming")
async def api_incoming_list(
    line: str = Query(None),
    search: str = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None),
    limit: int = Query(500),
):
    try:
        rows = incoming_query(line=line, search=search,
                              date_from=date_from, date_to=date_to, limit=limit)
        return JSONResponse({"ok": True, "records": rows, "count": len(rows)})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/item-lookup/{item_code:path}")
async def api_item_lookup(item_code: str):
    """ค้นหา description จาก item_code"""
    try:
        conn = sqlite3.connect(INCOMING_DB)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT item_code, description, part_type FROM item_descriptions WHERE item_code=?",
            (item_code.strip(),)
        ).fetchone()
        conn.close()
        if row:
            return JSONResponse({"ok": True, "item_code": row["item_code"],
                                 "description": row["description"], "part_type": row["part_type"]})
        return JSONResponse({"ok": False, "description": None})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})

@app.get("/api/vendors")
async def api_vendors_list():
    try:
        conn = sqlite3.connect(INCOMING_DB)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT vendor_code, vendor_name FROM vendors ORDER BY vendor_code").fetchall()
        conn.close()
        return JSONResponse({"ok": True, "vendors": [dict(r) for r in rows]})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})

@app.post("/api/vendors")
async def api_vendors_upsert(request: Request):
    try:
        body = await request.json()
        code = str(body.get("vendor_code","")).strip()
        name = str(body.get("vendor_name","")).strip()
        if not code or not name:
            return JSONResponse({"ok": False, "error": "vendor_code and vendor_name required"}, status_code=400)
        conn = sqlite3.connect(INCOMING_DB)
        conn.execute(
            "INSERT INTO vendors (vendor_code, vendor_name, updated_at) VALUES (?,?,?) "
            "ON CONFLICT(vendor_code) DO UPDATE SET vendor_name=excluded.vendor_name, updated_at=excluded.updated_at",
            (code, name, datetime.now().isoformat())
        )
        conn.commit(); conn.close()
        return JSONResponse({"ok": True, "vendor_code": code, "vendor_name": name})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})

@app.delete("/api/vendors/{vendor_code}")
async def api_vendors_delete(vendor_code: str):
    try:
        conn = sqlite3.connect(INCOMING_DB)
        conn.execute("DELETE FROM vendors WHERE vendor_code=?", (vendor_code.strip(),))
        conn.commit(); conn.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})

@app.get("/api/incoming/stats")
async def api_incoming_stats():
    try:
        return JSONResponse({"ok": True, "stats": incoming_stats()})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/incoming/stock")
async def api_incoming_stock(line: str = Query(None)):
    """สรุป Stock คงเหลือ = รับเข้า (incoming.db) − จ่ายออก (mecp.db ที่ จัดและส่งแล้ว)"""
    try:
        db_abs = os.path.abspath(DB_FILE)
        conn = sqlite3.connect(INCOMING_DB)
        conn.row_factory = sqlite3.Row
        conn.execute(f"ATTACH DATABASE ? AS mecp", (db_abs,))

        # ── รับเข้า: รวมทุกไลน์ต่อ Part No ──
        q_in = """
            SELECT part_no, item_code, unit,
                   SUM(qty)        AS received_qty,
                   COUNT(*)        AS lot_count,
                   MAX(receive_date) AS last_receive
            FROM incoming
            WHERE part_no IS NOT NULL AND part_no != ''
        """
        q_in += " GROUP BY part_no, item_code, unit"
        in_rows = conn.execute(q_in).fetchall()

        # ── รับเข้า: หา first_receive ต่อ part_no ด้วย ──
        first_recv_rows = conn.execute("""
            SELECT part_no, MIN(receive_date) AS first_receive
            FROM incoming GROUP BY part_no
        """).fetchall()
        first_receive = {str(r["part_no"] or "").strip(): str(r["first_receive"] or "") for r in first_recv_rows}

        # ── จ่ายออก: mecp.db orders ที่ workStatus = 'จัดและส่งแล้ว' ──
        # รวมทุกไลน์ตาม Part No เดียวกัน
        out_rows = conn.execute("""
            SELECT item_code AS part_no,
                   CASE
                     WHEN COALESCE(qty_adjusted, 0) = 1 THEN qty
                     WHEN part = 'U-Bend' AND teem_qty > 0 THEN teem_qty
                     ELSE qty
                   END AS qty,
                   COALESCE(confirm_date, date) AS dispatch_date
            FROM mecp.orders
            WHERE work_status = 'จัดและส่งแล้ว'
              AND item_code IS NOT NULL AND item_code != ''
        """).fetchall()

        # ── หักถาวร: deductions table (ดึงก่อน close connection) ──
        deduct_rows = conn.execute(
            "SELECT part_no, SUM(qty) AS total FROM deductions GROUP BY part_no"
        ).fetchall()

        # ── item descriptions ──
        desc_rows = conn.execute(
            "SELECT item_code, description FROM item_descriptions"
        ).fetchall()

        conn.execute("DETACH DATABASE mecp")
        conn.close()

        desc_map = {str(r["item_code"] or "").strip(): str(r["description"] or "") for r in desc_rows}

        def parse_th_date(d):
            """DD/MM/YYYY → YYYY-MM-DD สำหรับเปรียบเทียบกับ receive_date"""
            try:
                p = d.strip().split("/")
                return f"{p[2]}-{p[1].zfill(2)}-{p[0].zfill(2)}"
            except Exception:
                return ""

        # key = part_no รวมทุกไลน์
        # ถ้า part มีใน incoming → นับเฉพาะ dispatch ที่เกิดหลัง first_receive
        # ถ้า part ไม่เคยรับเข้า → นับ dispatch ทั้งหมด (แสดง stock ติดลบ)
        dispatch_by_pn = {}
        for r in out_rows:
            pn = str(r["part_no"] or "").strip()
            fr_raw = first_receive.get(pn, "")
            if fr_raw:
                # มีประวัติรับเข้า → นับเฉพาะ dispatch หลัง first_receive
                fr    = parse_th_date(fr_raw)
                ddate = parse_th_date(str(r["dispatch_date"] or ""))
                if not (ddate and fr and ddate >= fr):
                    continue
            # ไม่มีใน incoming → นับทุก dispatch
            dispatch_by_pn[pn] = dispatch_by_pn.get(pn, 0) + int(r["qty"] or 0)

        # deductions รวมตาม Part No
        deduct_by_pn = {str(r["part_no"] or "").strip(): int(r["total"] or 0) for r in deduct_rows}

        result = []
        covered_pn = set()
        for r in in_rows:
            ic   = str(r["item_code"] or "").strip()
            pn   = str(r["part_no"]   or "").strip()
            recv = int(r["received_qty"] or 0)
            covered_pn.add(pn)
            disp = dispatch_by_pn.get(pn, 0) + deduct_by_pn.get(pn, 0)
            net  = recv - disp   # อาจติดลบได้ถ้าจ่ายออกมากกว่ารับเข้า
            result.append({
                "part_no":        r["part_no"]     or "",
                "item_code":      ic,
                "line_num":       "",
                "unit":           r["unit"]        or "PC",
                "received_qty":   recv,
                "dispatched_qty": disp,
                "net_qty":        net,
                "lot_count":      int(r["lot_count"] or 0),
                "last_receive":   r["last_receive"] or "",
                "description":    desc_map.get(pn, ""),
            })

        # เพิ่ม rows สำหรับ part ที่ dispatch แล้ว แต่ไม่เคยรับเข้าใน incoming
        for pn, disp_qty in dispatch_by_pn.items():
            if pn in covered_pn:
                continue
            disp = disp_qty + deduct_by_pn.get(pn, 0)
            result.append({
                "part_no":        pn,
                "item_code":      pn,
                "line_num":       "",
                "unit":           "PC",
                "received_qty":   0,
                "dispatched_qty": disp,
                "net_qty":        -disp,
                "lot_count":      0,
                "last_receive":   "",
                "description":    desc_map.get(pn, ""),
            })

        result.sort(key=lambda x: x["net_qty"], reverse=True)
        return JSONResponse({"ok": True, "stock": result})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.delete("/api/incoming/{rec_id}")
async def api_incoming_delete(rec_id: str):
    try:
        conn = sqlite3.connect(INCOMING_DB)
        conn.execute("DELETE FROM incoming WHERE id=?", (rec_id,))
        conn.commit(); conn.close()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.patch("/api/incoming/{rec_id}/qty")
async def api_incoming_update_qty(rec_id: str, request: Request):
    """แก้ไขจำนวน (qty) ของรายการรับ part"""
    try:
        body = await request.json()
        new_qty = int(body.get("qty"))
        if new_qty < 0:
            return JSONResponse({"ok": False, "error": "qty ต้องไม่ติดลบ"}, status_code=400)
        conn = sqlite3.connect(INCOMING_DB)
        cur = conn.execute("UPDATE incoming SET qty=? WHERE id=?", (new_qty, rec_id))
        conn.commit()
        changed = cur.rowcount
        conn.close()
        if not changed:
            return JSONResponse({"ok": False, "error": "ไม่พบรายการ"}, status_code=404)
        return JSONResponse({"ok": True, "qty": new_qty})
    except (TypeError, ValueError):
        return JSONResponse({"ok": False, "error": "qty ต้องเป็นตัวเลข"}, status_code=400)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.delete("/api/incoming")
async def api_incoming_delete_bulk(before: str = Query(None)):
    """ลบทั้งหมด หรือลบก่อนวันที่ (before=YYYY-MM-DD)"""
    try:
        conn = sqlite3.connect(INCOMING_DB)
        if before:
            # before เป็น YYYY-MM-DD → receive_date เก็บเป็น DD/MM/YYYY
            from datetime import date as _date
            cutoff = _date.fromisoformat(before)
            # แปลง receive_date ทุก row แล้วเปรียบเทียบ
            rows = conn.execute("SELECT id, receive_date FROM incoming").fetchall()
            del_ids = []
            for rid, rd in rows:
                try:
                    d, m, y = rd.split('/')
                    row_date = _date(int(y), int(m), int(d))
                    if row_date < cutoff:
                        del_ids.append(rid)
                except Exception:
                    pass
            for rid in del_ids:
                conn.execute("DELETE FROM incoming WHERE id=?", (rid,))
            n = len(del_ids)
        else:
            n = conn.execute("SELECT COUNT(*) FROM incoming").fetchone()[0]
            conn.execute("DELETE FROM incoming")
        conn.commit(); conn.close()
        return JSONResponse({"ok": True, "deleted": n})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ───────────────────────────────────────────────
# WebSocket Manager
# ───────────────────────────────────────────────
class ConnectionManager:
    def __init__(self):
        self.active: list[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self.active.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self.active:
            self.active.remove(ws)

    async def broadcast(self, payload: dict):
        msg = json.dumps(payload, ensure_ascii=False)
        dead = []
        for ws in self.active:
            try:
                await ws.send_text(msg)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)

    async def send_to(self, ws: WebSocket, payload: dict):
        try:
            await ws.send_text(json.dumps(payload, ensure_ascii=False))
        except Exception:
            self.disconnect(ws)

manager = ConnectionManager()

# ───────────────────────────────────────────────
# WebSocket Endpoint
# ───────────────────────────────────────────────
@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    global order_counter
    await manager.connect(ws)
    client = ws.client
    print(f"[WS] เชื่อมต่อ: {client.host} | รวม: {len(manager.active)}")
    await manager.send_to(ws, {"type": "INIT", "orders": orders, "counter": order_counter})

    try:
        while True:
            raw = await ws.receive_text()
            try:
                msg = json.loads(raw)
            except json.JSONDecodeError:
                continue

            msg_type = msg.get("type")
            data = msg.get("data", {})

            if msg_type == "CREATE_ORDER":
                part      = data.get("part")
                model     = data.get("model", "N/A")
                qty       = data.get("qty")
                priority  = data.get("priority", "normal")
                remark    = data.get("remark", "")
                lot       = data.get("lot")
                line_num  = data.get("lineNum")
                seq       = data.get("seq")
                pro_month = data.get("proMonth")
                employee  = data.get("employee")
                item_code = data.get("itemCode")
                item_cond = data.get("itemCond")   # U-Bend lookup key (after SMP in QR)
                raw_qr    = data.get("rawQr")
                beg_qty      = data.get("begQty")
                teem_qty     = data.get("teemQty")
                need_by      = data.get("needBy")
                machine      = data.get("machine")
                qty_adjusted = 1 if data.get("qtyAdjusted") else 0
                original_qty = data.get("originalQty")
                ubend_sub    = data.get("ubendSub")  # 'Main' | 'Sub' | None
                prod_order   = data.get("prodOrder")  # Production Order No

                if not part or not qty:
                    await manager.send_to(ws, {"type": "ERROR", "message": "ข้อมูลไม่ครบ"})
                    continue

                now = datetime.now()
                order_id = f"ORD-{order_counter:04d}"
                order_counter += 1
                order = {
                    "id": order_id, "part": part, "model": model or "N/A",
                    "qty": int(qty), "priority": priority, "remark": remark,
                    "time": now.strftime("%H:%M:%S"),
                    "date": now.strftime("%d/%m/%Y"),
                    "timestamp": now.isoformat(),
                    "status": "pending", "confirmTime": None, "workStatus": None,
                    "lot": lot, "lineNum": line_num, "seq": seq, "proMonth": pro_month,
                    "employee": employee, "itemCode": item_code, "itemCond": item_cond, "rawQr": raw_qr,
                    "begQty": beg_qty, "teemQty": teem_qty, "needBy": need_by, "machine": machine,
                    "qtyAdjusted": bool(qty_adjusted), "originalQty": original_qty,
                    "ubendSub": ubend_sub,
                    "prodOrder": prod_order,
                }
                orders.append(order)
                # trim ให้เหลือแค่ ORDERS_LIMIT รายการล่าสุด — ที่เหลืออยู่ใน SQLite
                if len(orders) > ORDERS_LIMIT:
                    orders.sort(key=lambda o: o.get("timestamp", ""), reverse=True)
                    orders[:] = orders[:ORDERS_LIMIT]
                save_json()
                db_insert(order)
                # auto-upsert description จาก raw_qr ใหม่
                if raw_qr and item_code:
                    _upsert_item_desc(item_code, _extract_desc_from_qr(raw_qr, item_code), part)
                await manager.broadcast({"type": "ORDER_CREATED", "order": order, "counter": order_counter})
                print(f"[ORDER] สร้าง: {order_id} | {part} | Qty:{qty} | LOT:{lot}")

            elif msg_type == "SELECT_ORDER":
                order_id = data.get("id")
                order = next((o for o in orders if o["id"] == order_id), None)
                if order and order["status"] != "confirmed":
                    await manager.broadcast({"type": "ORDER_SELECTED", "id": order_id})

            elif msg_type == "UPDATE_WORK_STATUS":
                order_id    = data.get("id")
                work_status = data.get("workStatus")
                print(f"[ORDER] UPDATE_WORK_STATUS: {order_id} -> {work_status}")
                order = next((o for o in orders if o["id"] == order_id), None)
                if order:
                    order["workStatus"] = work_status
                    save_json()
                    try:
                        conn = sqlite3.connect(DB_FILE)
                        conn.execute("UPDATE orders SET work_status=? WHERE id=?", (work_status, order_id))
                        conn.commit()
                        conn.close()
                    except Exception as e:
                        print(f"[DB ERROR] UPDATE_WORK_STATUS: {e}")
                    await manager.broadcast({"type": "ORDER_UPDATED", "order": order})
                    print(f"[ORDER] Broadcasted ORDER_UPDATED: {order_id}")
                else:
                    print(f"[ORDER] UPDATE_WORK_STATUS: order {order_id} not found")

            elif msg_type == "CONFIRM_ORDER":
                order_id    = data.get("id")
                work_status = data.get("workStatus", "จัดและส่งแล้ว")
                confirm_by    = data.get("confirmBy")
                floor_remark  = data.get("floorRemark")
                rep_line       = (data.get("repLine") or "").strip()
                rep_seq        = (data.get("repSeq") or "").strip()
                rep_prod_month = (data.get("repProdMonth") or "").strip()
                order = next((o for o in orders if o["id"] == order_id), None)
                if not order:
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} ไม่พบ"})
                    continue
                if order["status"] == "confirmed":
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} Confirm แล้ว"})
                    continue
                now_dt = datetime.now()
                ct = now_dt.strftime("%H:%M:%S")
                cd = now_dt.strftime("%d/%m/%Y")
                # Order ทดแทน: ต้องระบุ LINE/SEQ/Prod Month ของ part ที่นำมาทดแทน
                is_replace = str(order.get("remark") or "").startswith("[ทดแทน:")
                if is_replace:
                    if not (rep_line and rep_seq and rep_prod_month):
                        await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} Order ทดแทนต้องระบุ LINE/SEQ/Prod Month"})
                        continue
                    sub_txt = f"[ทดแทนด้วย: LINE {rep_line} / SEQ {rep_seq} / Prod {rep_prod_month}]"
                    floor_remark = f"{floor_remark} {sub_txt}".strip() if floor_remark else sub_txt
                    order["replaceSub"] = {"line": rep_line, "seq": rep_seq, "prodMonth": rep_prod_month}
                order["status"] = "confirmed"
                order["workStatus"] = work_status
                order["confirmTime"] = ct
                order["confirmDate"] = cd
                order["confirmBy"]    = confirm_by
                order["floorRemark"] = floor_remark
                save_json()
                db_confirm(order_id, work_status, ct, confirm_by, cd, floor_remark)
                await manager.broadcast({"type": "ORDER_CONFIRMED", "order": order})
                print(f"[ORDER] Confirmed: {order_id} by {confirm_by}")

            elif msg_type == "UNCONFIRM_ORDER":
                order_id = data.get("id")
                order = next((o for o in orders if o["id"] == order_id), None)
                if not order:
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} ไม่พบ"})
                    continue
                if order["status"] != "confirmed":
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} ยังไม่ได้ Confirm"})
                    continue
                order["status"] = "pending"
                order["workStatus"] = None
                order["confirmTime"] = None
                order["confirmDate"] = None
                order["confirmBy"] = None
                order["floorRemark"] = None
                order.pop("replaceSub", None)
                save_json()
                db_unconfirm(order_id)
                await manager.broadcast({"type": "ORDER_UNCONFIRMED", "order": order})
                print(f"[ORDER] Unconfirmed: {order_id}")

            elif msg_type == "RECEIVE_ORDER":
                order_id   = data.get("id")
                receive_by = data.get("receiveBy")
                order = next((o for o in orders if o["id"] == order_id), None)
                if not order:
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} ไม่พบ"})
                    continue
                if order.get("workStatus") != "จัดและส่งแล้ว":
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} ยังไม่ได้จัดและส่ง"})
                    continue
                now_dt = datetime.now()
                rt = now_dt.strftime("%H:%M:%S")
                rd = now_dt.strftime("%d/%m/%Y")
                order["workStatus"]   = "รับแล้ว"
                order["receiveTime"]  = rt
                order["receiveDate"]  = rd
                order["receiveBy"]    = receive_by
                save_json()
                db_receive(order_id, rt, rd, receive_by)
                await manager.broadcast({"type": "ORDER_RECEIVED", "order": order})
                print(f"[ORDER] Received: {order_id} by {receive_by}")

            elif msg_type == "CANCEL_ORDER":
                order_id = data.get("id")
                idx = next((i for i, o in enumerate(orders) if o["id"] == order_id), None)
                if idx is None:
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} ไม่พบ"})
                    continue
                order = orders[idx]
                if order.get("status") == "confirmed":
                    await manager.send_to(ws, {"type": "ERROR", "message": f"{order_id} ถูก Confirm แล้ว ยกเลิกไม่ได้"})
                    continue
                orders.pop(idx)
                save_json()
                conn = sqlite3.connect(DB_FILE)
                conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
                conn.commit()
                conn.close()
                await manager.broadcast({"type": "ORDER_CANCELLED", "id": order_id})
                print(f"[ORDER] Cancelled: {order_id}")

            elif msg_type == "PING":
                await manager.send_to(ws, {"type": "PONG"})

            elif msg_type == "GET_STATE":
                await manager.send_to(ws, {"type": "INIT", "orders": orders, "counter": order_counter})

    except WebSocketDisconnect:
        manager.disconnect(ws)
        print(f"[WS] ตัดการเชื่อมต่อ: {client.host} | รวม: {len(manager.active)}")

# ───────────────────────────────────────────────
# REST API — Pages
# ───────────────────────────────────────────────
@app.get("/")
async def root():
    from fastapi.responses import FileResponse
    return FileResponse("public/index.html")

@app.get("/dashboard")
async def dashboard():
    from fastapi.responses import FileResponse
    return FileResponse("public/dashboard.html")

@app.get("/db")
async def db_viewer():
    from fastapi.responses import FileResponse
    return FileResponse("public/db.html")

@app.get("/ubend-editor")
async def ubend_editor():
    from fastapi.responses import FileResponse
    return FileResponse("public/ubend-editor.html")

# ───────────────────────────────────────────────
# REST API — SQLite endpoints
# ───────────────────────────────────────────────
@app.delete("/api/db/orders/{order_id}")
async def api_db_delete_one(order_id: str, stock_action: str = Query("return")):
    """ลบ order เดียว
    stock_action='keep'   → บันทึก deduction เพื่อคง stock
    stock_action='return' → ลบแล้วคืนจำนวนกลับ stock (default)
    """
    global orders
    if stock_action == "keep":
        order = next((o for o in orders if o["id"] == order_id), None) or db_get_order(order_id)
        if order and (order.get("workStatus") or order.get("work_status")) == "จัดและส่งแล้ว":
            create_deduction(order)
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
        conn.commit()
        conn.close()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    orders = [o for o in orders if o["id"] != order_id]
    save_json()
    await manager.broadcast({"type": "INIT", "orders": orders, "counter": order_counter})
    return JSONResponse({"message": f"Deleted {order_id}"})

@app.delete("/api/db/orders")
async def api_db_delete_all(before: str = Query(None), stock_action: str = Query("return")):
    """ลบหลาย order (ทั้งหมด หรือ ก่อนวันที่)
    stock_action='keep'   → บันทึก deduction สำหรับทุก order ที่ 'จัดและส่งแล้ว'
    stock_action='return' → ลบแล้วคืนจำนวนกลับ stock (default)
    """
    global orders, order_counter
    # หา orders ที่จะถูกลบก่อน (เพื่อสร้าง deductions)
    if stock_action == "keep":
        to_delete = db_orders_for_delete(before)
        for o in to_delete:
            if (o.get("workStatus") or o.get("work_status")) == "จัดและส่งแล้ว":
                create_deduction(o)
    try:
        conn = sqlite3.connect(DB_FILE)
        if before:
            del_ids = [o.get("id") for o in db_orders_for_delete(before) if o.get("id")]
            conn.executemany("DELETE FROM orders WHERE id=?", [(oid,) for oid in del_ids])
        else:
            conn.execute("DELETE FROM orders")
        conn.commit()
        conn.close()
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
    if before:
        orders = [o for o in orders if _th_date_to_iso(o.get("date","")) >= before]
    else:
        orders = []
        order_counter = 1
    save_json()
    await manager.broadcast({"type": "INIT", "orders": orders, "counter": order_counter})
    return JSONResponse({"message": "Deleted"})

@app.get("/api/db/orders")
async def api_db_orders(
    status:   str = Query(None),
    part:     str = Query(None),
    priority: str = Query(None),
    search:   str = Query(None),
    limit:    int = Query(1000),
):
    rows = db_query(status, part, priority, search, limit)
    return JSONResponse({"orders": rows, "total": len(rows)})

@app.get("/api/db/stats")
async def api_db_stats():
    return JSONResponse(db_stats())

# ───────────────────────────────────────────────
# REST API — JSON (เดิม)
# ───────────────────────────────────────────────
@app.get("/api/orders")
async def get_orders():
    return JSONResponse({"orders": orders, "total": len(orders)})

@app.delete("/api/orders")
async def clear_orders():
    global orders, order_counter
    orders = []
    order_counter = 1
    save_json()
    await manager.broadcast({"type": "INIT", "orders": [], "counter": 1})
    return JSONResponse({"message": "Orders cleared"})

# ───────────────────────────────────────────────
# ── CUTTING CHECK ─────────────────────────────────────────────────────────────
_CUT_DIR = os.path.dirname(__file__)
_CUT_DB  = os.path.join(_CUT_DIR, "cutting.db")

def cutting_get_db():
    con = sqlite3.connect(_CUT_DB)
    con.row_factory = sqlite3.Row
    return con

def cutting_init_db():
    con = sqlite3.connect(_CUT_DB)
    con.execute('''CREATE TABLE IF NOT EXISTS cutting_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        saved_at TEXT, sheet_no TEXT,
        id_card TEXT, mc_line TEXT, tag_fg_no1 TEXT, tag_fg_no2 TEXT,
        item_insulation TEXT,
        check_laydown TEXT, check_burr TEXT, check_drawing TEXT,
        check_copper TEXT, check_apprence TEXT,
        item_dwg TEXT, item_type TEXT, line TEXT, item_name TEXT,
        seq TEXT, lot TEXT, prod_month TEXT, mc_cutting TEXT,
        insulation_qty TEXT, cutting_pcs TEXT,
        lay_down_side TEXT, lay_down_middle TEXT,
        process_scan TEXT, leader_confirm TEXT, status TEXT
    )''')
    con.commit()
    con.close()

cutting_init_db()

@app.get("/cutting")
@app.get("/cutting/")
async def cutting_page():
    return FileResponse(os.path.join(_CUT_DIR, "cutting.html"))

@app.get("/cutting/data/{filename:path}")
async def cutting_data(filename: str):
    path = os.path.join(_CUT_DIR, "data", filename)
    if os.path.exists(path):
        return FileResponse(path)
    return JSONResponse({"error": "not found"}, status_code=404)

@app.get("/cutting/images/{filename:path}")
async def cutting_images_route(filename: str):
    path = os.path.join(_CUT_DIR, "images", filename)
    if os.path.exists(path):
        return FileResponse(path)
    return JSONResponse({"error": "not found"}, status_code=404)

@app.get("/cutting/api/cutting-records")
async def cutting_records_api():
    con = cutting_get_db()
    rows = [dict(r) for r in con.execute(
        "SELECT * FROM cutting_records ORDER BY id DESC").fetchall()]
    con.close()
    return JSONResponse(rows)

@app.post("/cutting/api/cutting-record")
async def cutting_save_api(request: Request):
    d = await request.json()
    try:
        con = cutting_get_db()
        con.execute('''
            INSERT INTO cutting_records
            (saved_at, sheet_no,
             id_card, mc_line, tag_fg_no1, tag_fg_no2, item_insulation,
             check_laydown, check_burr, check_drawing, check_copper, check_apprence,
             item_dwg, item_type, line, item_name, seq, lot,
             prod_month, mc_cutting, insulation_qty, cutting_pcs,
             lay_down_side, lay_down_middle,
             process_scan, leader_confirm, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            datetime.now().isoformat(), d.get('sheetNo','CH-541-11-12-01'),
            d.get('idCard'),       d.get('mcLine'),
            d.get('tagFgNo1'),     d.get('tagFgNo2'),    d.get('itemInsulation'),
            d.get('checkLaydown'), d.get('checkBurr'),   d.get('checkDrawing'),
            d.get('checkCopper'),  d.get('checkApprence'),
            d.get('itemDwg'),      d.get('type'),         d.get('line'),
            d.get('itemName'),     d.get('seq'),          d.get('lot'),
            d.get('prodMonth'),    d.get('mcCutting'),
            d.get('insulationQty'), d.get('cuttingPcs'),
            d.get('layDownSide'),  d.get('layDownMiddle'),
            d.get('processScan'),  d.get('leaderConfirm'), d.get('status'),
        ))
        con.commit()
        rid = con.execute('SELECT last_insert_rowid()').fetchone()[0]
        con.close()
        return JSONResponse({'ok': True, 'id': rid})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/cutting/process-control/{filename:path}")
async def cutting_process_control(filename: str):
    _PC_DIR = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\JPEG FILE"
    for fn in [filename, filename.replace('.jpg','.JPG'), filename.replace('.JPG','.jpg')]:
        path = os.path.join(_PC_DIR, fn)
        if os.path.exists(path):
            return FileResponse(path)
    return JSONResponse({"error": "not found", "file": filename}, status_code=404)

@app.put("/cutting/api/cutting-record/{rid}")
async def cutting_update_api(rid: int, request: Request):
    d = await request.json()
    allowed = ['id_card','mc_line','tag_fg_no1','tag_fg_no2','item_insulation',
               'check_laydown','check_burr','check_drawing','check_copper','check_apprence',
               'item_dwg','item_type','line','item_name','seq','lot',
               'prod_month','mc_cutting','insulation_qty','cutting_pcs',
               'lay_down_side','lay_down_middle','leader_confirm','status']
    sets = ', '.join(f'{c}=?' for c in allowed)
    vals = [d.get(c,'') for c in allowed] + [rid]
    try:
        con = cutting_get_db()
        con.execute(f'UPDATE cutting_records SET {sets} WHERE id=?', vals)
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.delete("/cutting/api/cutting-record/{rid}")
async def cutting_delete_api(rid: int):
    try:
        con = cutting_get_db()
        con.execute('DELETE FROM cutting_records WHERE id=?', (rid,))
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/cutting/api/cutting-export-excel")
async def cutting_export_excel_api():
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from fastapi.responses import StreamingResponse
        con = cutting_get_db()
        rows = [dict(r) for r in con.execute(
            "SELECT * FROM cutting_records ORDER BY id").fetchall()]
        con.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cutting Records'
        headers = ['No','วันที่-เวลา','ID Card','MC/Line','Tag FG No.1','Tag FG No.2',
                   'Item Insulation','Type','Line','Lay Down','Burr','Drawing','Copper','Apprence',
                   'Insul.Qty','Cut.Pcs','Lay Side','Lay Mid','Leader','Status']
        ws.append(headers)
        thin = Side(style='thin', color='B0BEC5')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='263238')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center')
        ok_fill = PatternFill('solid', fgColor='C8E6C9')
        ng_fill = PatternFill('solid', fgColor='FFCDD2')
        for i, r in enumerate(rows, 1):
            ws.append([
                i, r.get('saved_at',''), r.get('id_card',''), r.get('mc_line',''),
                r.get('tag_fg_no1',''), r.get('tag_fg_no2',''), r.get('item_insulation',''),
                r.get('item_type',''), r.get('line',''),
                r.get('check_laydown',''), r.get('check_burr',''), r.get('check_drawing',''),
                r.get('check_copper',''), r.get('check_apprence',''),
                r.get('insulation_qty',''), r.get('cutting_pcs',''),
                '✓' if r.get('lay_down_side') else '',
                '✓' if r.get('lay_down_middle') else '',
                r.get('leader_confirm',''), r.get('status',''),
            ])
            for cell in ws[i+1]: cell.border = bdr
            st = str(r.get('status') or '')
            ws.cell(i+1, len(headers)).fill = (
                ok_fill if st == 'OK' else (ng_fill if st == 'NG' else PatternFill()))
        col_widths = [5,18,20,22,16,16,16,10,8,10,8,10,8,10,10,10,10,10,20,8]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = ws.dimensions
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"cutting_log_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

# ───────────────────────────────────────────────
# Start
# ───────────────────────────────────────────────
if __name__ == "__main__":
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80)); local_ip = s.getsockname()[0]
    except Exception:
        local_ip = "localhost"
    finally:
        s.close()

    print("")
    print("╔══════════════════════════════════════════════════╗")
    print("║   MECP + Cutting — SERVER :3000       ║")
    print("╠══════════════════════════════════════════════════╣")
    print(f"║  Network IP   : http://{local_ip}:3000          ║")
    print(f"║  Main         : /                                ║")
    print(f"║  Dashboard    : /dashboard                       ║")
    print(f"║  DB Viewer    : /db                              ║")
    print(f"║  DB: mecp.db  : orders (planner/floor)           ║")
    print(f"║  DB: incoming : รับ Part (receiver)              ║")
    print(f"║  ✂ Cutting   : /cutting                         ║")
    print("╚══════════════════════════════════════════════════╝")
    print("")

    uvicorn.run("server:app", host="0.0.0.0", port=3000, reload=False)

