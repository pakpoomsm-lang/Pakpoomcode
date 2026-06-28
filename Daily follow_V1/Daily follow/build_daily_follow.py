import json
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, time
from html import escape
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
EXPORT_FILE = ROOT / "EXPORT_20260601070215.xlsx"
SAMPLE_FILE = ROOT / "HEI-18-05-PP-Original-SAP-1 (1).xlsm"
ROUTING_FILE = ROOT / "Routing.xlsx"
MASTER_TS_FILE = ROOT / "MasterTS.xlsx"
MASTER_TS_DB_FILE = ROOT / "MasterTS.db"
PROGRESS_FILE = ROOT / "ZPP0059.xlsx"
CT_DB_FILE = ROOT / "CT.db"
CT_XLSX_FILE = ROOT / "CT.xlsx"
OUTPUT_FILE = ROOT / "daily_follow.html"
TEMPLATE_FILE = ROOT / "daily_follow_template.html"

# Map ZPP0059 "Operation Short Text" -> progress column (material-level)
OP_TO_FIELD = {"Insert": "fp", "Brazing": "auto", "Cutting": "cutting"}
# H/P bender is recorded on sub-component materials, so it is summed per assembly lot.
HP_OP = "H/P bender"


def norm(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value):
    if value in (None, ""):
        return ""
    try:
        f = float(str(value).strip())
    except ValueError:
        return value
    return int(f) if f.is_integer() else round(f, 3)


def ts_number(value):
    """Like number() but keeps 6 decimals. TS values can have 4 decimals
    (e.g. 0.0485); rounding to 3 like number() would collapse it to 0.049 and
    skew the unit rate (0.0485 -> 16.4, but 0.049 -> 16.6 for a lot of 200)."""
    if value in (None, ""):
        return ""
    try:
        f = float(str(value).strip())
    except ValueError:
        return value
    return int(f) if f.is_integer() else round(f, 6)


def excel_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    if " " in raw:
        raw = raw.split(" ", 1)[0]
    try:
        return datetime.fromisoformat(raw).date().isoformat()
    except ValueError:
        return raw


def excel_time(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    raw = str(value).strip()
    if " " in raw:
        raw = raw.split(" ", 1)[-1]
    parts = raw.split(":")
    if len(parts) >= 2:
        return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    return raw


def parse_dt(d_value, t_value):
    d = excel_date(d_value)
    t = excel_time(t_value) or "00:00"
    if not d:
        return None
    try:
        return datetime.fromisoformat(f"{d}T{t}")
    except ValueError:
        return None


def month_display(value):
    raw = text(value)
    if "." in raw:
        left, right = raw.split(".", 1)
        try:
            return f"{int(left)}.{right}"
        except ValueError:
            return raw
    return raw


def _db_needs_build(db_path, xlsx_path):
    """True if the SQLite db should be (re)built from the xlsx: db missing, or
    the xlsx has been edited since the db was built. Without this a stale db
    would silently shadow an updated spreadsheet."""
    if not xlsx_path.exists():
        return False  # no source to build from; use the db as-is (or none)
    if not db_path.exists():
        return True
    return xlsx_path.stat().st_mtime > db_path.stat().st_mtime


def load_master_ts():
    """Return {ITEM: TS_VALUE} from MasterTS.db (built from MasterTS.xlsx)."""
    if _db_needs_build(MASTER_TS_DB_FILE, MASTER_TS_FILE):
        import build_masterts_db  # (re)build MasterTS.db from the xlsx
        build_masterts_db.main()
    if not MASTER_TS_DB_FILE.exists():
        return {}
    con = sqlite3.connect(MASTER_TS_DB_FILE)
    rows = con.execute(
        "SELECT item, ts_value FROM master_ts WHERE item IS NOT NULL AND item <> ''"
    ).fetchall()
    con.close()
    return {item: ts_number(ts) for item, ts in rows}


def load_routing():
    master_ts = load_master_ts()
    wb = openpyxl.load_workbook(ROUTING_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    routing = {}
    for row in rows:
        material = text(row[idx["Material"]])
        if not material or material in routing:
            continue
        ts = master_ts.get(material, ts_number(row[idx["OPR Time Standard"]]))
        routing[material] = {
            "description": text(row[idx["Description"]]),
            "operation": text(row[idx["Operation Text"]]),
            "attribute1": text(row[idx["Attribute1"]]),
            "attribute2": text(row[idx["Attribute2"]]),
            "unit": number(row[idx["Standard lot"]]),
            "ts": ts,
        }
    wb.close()
    return routing


def load_ct():
    """Return {material: {finpress, hpbender, expander, autoBrazing, cutting}}.

    Per-process cycle time in seconds, from CT.db (built by build_ct_db.py).
    A material can appear under several sub-types with different CTs, so we take
    the MAX of each process across them (worst-case capacity planning). Finpress
    -> Auto-Brazing use the *-CT columns; Cutting uses Cutting-Sec.
    """
    if _db_needs_build(CT_DB_FILE, CT_XLSX_FILE):
        import build_ct_db  # (re)build CT.db from CT.xlsx
        build_ct_db.main()
    if not CT_DB_FILE.exists():
        return {}
    con = sqlite3.connect(CT_DB_FILE)
    rows = con.execute(
        """
        SELECT material,
               MAX(finpress_ct), MAX(hpbender_ct), MAX(expander_ct),
               MAX(auto_brazing_ct), MAX(cutting_sec)
        FROM cycle_time
        WHERE material IS NOT NULL AND material <> ''
        GROUP BY material
        """
    ).fetchall()
    con.close()
    return {
        mat: {
            "finpress": fp or 0,
            "hpbender": hp or 0,
            "expander": ex or 0,
            "autoBrazing": ab or 0,
            "cutting": cut or 0,
        }
        for mat, fp, hp, ex, ab, cut in rows
    }


def seq_key(value):
    n = number(value)
    return str(n) if n != "" else ""


def final_qty(attribute, matp):
    """Finished quantity = qty of the last operation, decided by Attribute suffix.

    -BZ -> ends at Brazing (matp["auto"]); -CT -> ends at Cutting (matp["cutting"]).
    Anything else -> blank.
    """
    a = (attribute or "").strip().upper()
    if a.endswith("-BZ"):
        return matp.get("auto", "")
    if a.endswith("-CT"):
        return matp.get("cutting", "")
    return ""


def clean_qty(value):
    if not isinstance(value, (int, float)):
        return 0
    return int(value) if float(value).is_integer() else round(float(value), 2)


def load_progress():
    """Aggregate real per-process progress from ZPP0059.

    Returns (by_mat, by_lot):
      by_mat["line|month|seq|material"] = {"fp": q, "auto": q, "cutting": q}
      by_lot["line|month|seq|assyOrder"] = q   (H/P bender, summed per lot)
    """
    by_mat = defaultdict(lambda: defaultdict(float))
    by_lot = defaultdict(float)
    if not PROGRESS_FILE.exists():
        return {}, {}
    wb = openpyxl.load_workbook(PROGRESS_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    c_line = idx["Production Line"]
    c_month = idx["Production Month"]
    c_seq = idx["Sequence"]
    c_mat = idx["Material"]
    c_ao = idx["Assembly Order"]
    c_op = idx["Operation Short Text"]
    c_qty = idx["Posted Quantity"]
    # Optional in older exports — used to skip cancelled / deleted postings.
    c_del = idx.get("Deletion Flag")
    c_cancel = idx.get("Cancel Date")
    for row in rows:
        line = text(row[c_line])
        if not line:
            continue
        if c_del is not None and text(row[c_del]):
            continue
        if c_cancel is not None and text(row[c_cancel]):
            continue
        key_head = f"{line}|{month_display(row[c_month])}|{seq_key(row[c_seq])}"
        op = text(row[c_op])
        qty = row[c_qty] if isinstance(row[c_qty], (int, float)) else 0
        if op == HP_OP:
            by_lot[f"{key_head}|{text(row[c_ao])}"] += qty
        elif op in OP_TO_FIELD:
            by_mat[f"{key_head}|{text(row[c_mat])}"][OP_TO_FIELD[op]] += qty
    wb.close()
    mat_out = {
        k: {f: clean_qty(q) for f, q in fields.items()}
        for k, fields in by_mat.items()
    }
    lot_out = {k: clean_qty(q) for k, q in by_lot.items()}
    return mat_out, lot_out


def most_common(counter, default=""):
    if not counter:
        return default
    return counter.most_common(1)[0][0]


def load_speed_maps():
    by_line = defaultdict(Counter)
    by_model = defaultdict(Counter)
    wb = openpyxl.load_workbook(SAMPLE_FILE, read_only=True, data_only=True, keep_vba=False)
    for sheet in wb.sheetnames:
        if not sheet.startswith("Line "):
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=7, values_only=True):
            line = text(row[0]) if len(row) > 0 else ""
            code = text(row[3]) if len(row) > 3 else ""
            model = text(row[4]) if len(row) > 4 else ""
            speed = row[11] if len(row) > 11 else None
            if not line or not isinstance(speed, (int, float)):
                continue
            speed = int(speed) if float(speed).is_integer() else round(float(speed), 2)
            by_line[line][speed] += 1
            if code and model:
                by_model[(line, code, model)][speed] += 1
    wb.close()
    return (
        {line: most_common(counter) for line, counter in by_line.items()},
        {key: most_common(counter) for key, counter in by_model.items()},
    )


def assign_speed(rows):
    """Set speed as seconds-per-unit for each row:

        speed = (start of next seq - start of this seq) / assy_order

    The gap is measured per line within a month, ordering sequences by their
    production time. Every row sharing the same seq uses that seq's start time;
    the divisor is the row's own assy order quantity.
    """
    # Earliest production datetime per (line, month, seq).
    seq_start = {}
    for r in rows:
        if not r["prodSort"]:
            continue
        key = (r["line"], r["month"], r["seq"])
        dt = r["prodSort"]
        if key not in seq_start or dt < seq_start[key]:
            seq_start[key] = dt

    # Gap in seconds from each seq to the chronologically next seq on the line.
    gap_by_seq = {}
    by_line_month = defaultdict(list)
    for (line, month, seq), dt in seq_start.items():
        by_line_month[(line, month)].append((dt, seq))
    for (line, month), items in by_line_month.items():
        items.sort()
        for i in range(len(items) - 1):
            cur_dt, seq = items[i]
            nxt_dt = items[i + 1][0]
            gap = (
                datetime.fromisoformat(nxt_dt) - datetime.fromisoformat(cur_dt)
            ).total_seconds()
            gap_by_seq[(line, month, seq)] = gap

    # Assign sec/unit per row.
    for r in rows:
        gap = gap_by_seq.get((r["line"], r["month"], r["seq"]))
        assy = r["assyOrder"]
        if gap and gap > 0 and isinstance(assy, (int, float)) and assy > 0:
            r["speed"] = round(gap / assy, 2)
        else:
            r["speed"] = ""


# Logical column name -> (header text in the ZPP0022 export, legacy fixed index).
# The legacy index is only used as a fallback when the header is missing, so old
# exports keep working even if SAP drops/renames a column.
COLUMN_HEADERS = {
    "line": ("Line", 0),
    "seq": ("Sequence", 1),
    "month": ("Production Month", 2),
    "assyMaterial": ("Assy Material", 3),
    "assyMaterialDesc": ("Assy Material Description", 4),
    "orderQty": ("Order Quantity", 5),
    "material": ("Material", 6),
    "description": ("Description", 7),
    "status": ("Status", 8),
    "prodDate": ("Assy unit start date", 9),
    "prodTime": ("Assy unit start time", 10),
    "productionOrder": ("Production Order", 11),
    "assyOrderQty": ("Assy Order Quantity", 12),
    "assyOrder": ("Assy Order", 13),
    "remarkFeeder": ("Remark Order Feeder", 23),
    "remarkAssy": ("Remark Order Assy", 24),
    "assyStatus": ("Assy Status", 26),
    "metal": ("METAL", 27),
    "paint2": ("PAINT2", 28),
    "pipe1": ("PIPE1", 29),
    "pipe2": ("PIPE2", 30),
    "hex1out": ("HEX1OUT", 31),
    "hex1in": ("HEX1IN", 32),
    "hex2": ("HEX2", 33),
    "dirOrder": ("DIR.ORDER", 39),
    "dirGr": ("DIR.GR", 40),
    "dirRemark": ("DIR.REMARK", 41),
    "invRmk": ("INV.RMK", 44),
    "feederFinishTime": ("feeder finish time", 45),
    "lineRemark": ("Line Remark", 50),
    "schedFinishDate": ("Scheduled Finish Date", 51),
    "feederFinishDate": ("feeder finish date", 52),
    "schedFinishTime": ("Scheduled Finish Time", 53),
}

# Columns that a valid ZPP0022 order export must contain. If any is missing we
# stop instead of silently importing partial/wrong data. This also rejects the
# stripped SAP layout (starts with "Production Order", ~18 columns) which lacks
# Assy Material / Assy Order / Assy unit start date etc. that the dashboard needs.
REQUIRED_COLUMNS = (
    "line", "status", "material", "month", "seq",
    "assyMaterial", "assyMaterialDesc", "prodDate", "assyOrder", "assyStatus",
)


def resolve_columns(header_row):
    """Map each logical column to its index using the export's header row.

    Falls back to the legacy fixed index when a header is not found, so that
    older exports keep working. Raises if a REQUIRED column cannot be located.
    """
    lookup = {}
    for i, value in enumerate(header_row):
        if value is None:
            continue
        key = str(value).strip().lower()
        # First occurrence wins (the export has two "Description" columns).
        lookup.setdefault(key, i)

    cols, missing = {}, []
    for key, (name, fallback) in COLUMN_HEADERS.items():
        idx = lookup.get(name.strip().lower())
        if idx is None:
            idx = fallback
            missing.append(name)
        cols[key] = idx

    missing_required = [
        COLUMN_HEADERS[k][0] for k in REQUIRED_COLUMNS if COLUMN_HEADERS[k][0] in missing
    ]
    if missing_required:
        raise ValueError(
            "Export is missing required column(s): "
            + ", ".join(missing_required)
            + ". This does not look like a ZPP0022 order export."
        )
    if missing:
        print(
            "[warn] columns located by position (header not found): "
            + ", ".join(missing),
            file=sys.stderr,
        )
    return cols


def build_rows():
    routing = load_routing()
    prog_mat, prog_lot = load_progress()
    wb = openpyxl.load_workbook(EXPORT_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    col = resolve_columns(next(rows))
    out = []
    for row in rows:
        line = text(row[col["line"]])
        if not line:
            continue

        item = text(row[col["material"]])
        route = routing.get(item, {})
        prod_dt = parse_dt(row[col["prodDate"]], row[col["prodTime"]])
        finish_dt = parse_dt(
            row[col["schedFinishDate"]] or row[col["feederFinishDate"]],
            row[col["schedFinishTime"]] or row[col["feederFinishTime"]],
        )
        lead = ""
        if prod_dt and finish_dt:
            lead = round((prod_dt - finish_dt).total_seconds() / 3600, 2)

        assy_qty = number(row[col["assyOrderQty"]])
        status = text(row[col["status"]])
        code = text(row[col["assyMaterial"]])
        model = text(row[col["assyMaterialDesc"]])
        month = month_display(row[col["month"]])
        assy_order_no = text(row[col["assyOrder"]])
        head = f"{line}|{month}|{seq_key(row[col['seq']])}"
        matp = prog_mat.get(f"{head}|{item}", {})
        hp_val = prog_lot.get(f"{head}|{assy_order_no}", "")
        fin = final_qty(route.get("attribute2"), matp)
        remark = " ".join(
            part
            for part in [
                text(row[col["remarkFeeder"]]),
                text(row[col["remarkAssy"]]),
                text(row[col["lineRemark"]]),
            ]
            if part
        )

        out.append(
            {
                "line": line,
                "seq": number(row[1]),
                "month": month,
                "code": code,
                "model": model,
                "finished": "",
                "orderQty": number(row[col["orderQty"]]),
                "item": item,
                "description": route.get("description") or text(row[col["description"]]),
                "attribute": route.get("attribute2") or "",
                "attribute1": route.get("attribute1") or "",
                "productionOrder": text(row[col["productionOrder"]]),
                "speed": "",  # filled by assign_speed() after all rows are built
                "prodDate": excel_date(row[col["prodDate"]]),
                "prodTime": excel_time(row[col["prodTime"]]),
                "prodSort": prod_dt.isoformat() if prod_dt else "",
                "assyOrder": assy_qty,
                "assyOrderNo": assy_order_no,
                "status": status,
                "assyStatus": text(row[col["assyStatus"]]),
                "remark": remark,
                "hp": hp_val,
                "fp": matp.get("fp", ""),
                "exp": "",
                "auto": matp.get("auto", ""),
                "cutting": matp.get("cutting", ""),
                "fg": fin,
                "unit": route.get("unit") or "",
                "stockFg": fin,
                "subcooler": fin,
                "lead": lead,
                "leadRemark": text(row[col["dirRemark"]]) or text(row[col["invRmk"]]),
                "ts": route.get("ts") or "",
                "operation": route.get("operation") or "",
                "sourceReady": {
                    "metal": norm(row[col["metal"]]),
                    "paint2": norm(row[col["paint2"]]),
                    "pipe1": norm(row[col["pipe1"]]),
                    "pipe2": norm(row[col["pipe2"]]),
                    "hex1out": norm(row[col["hex1out"]]),
                    "hex1in": norm(row[col["hex1in"]]),
                    "hex2": norm(row[col["hex2"]]),
                    "dirOrder": norm(row[col["dirOrder"]]),
                    "dirGr": norm(row[col["dirGr"]]),
                },
            }
        )
    wb.close()
    assign_speed(out)
    out.sort(
        key=lambda r: (
            r["line"],
            r["prodSort"] or "9999",
            float(r["seq"]) if isinstance(r["seq"], (int, float)) else 999999,
            r["assyOrderNo"],
            r["productionOrder"],
            r["item"],
        )
    )
    return out


def render_html(rows):
    default_line = "ALL"
    default_date = ""
    routing = load_routing()
    speed_by_line, speed_by_model = load_speed_maps()
    prog_mat, prog_lot = load_progress()
    payload = json.dumps(
        {
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "sourceFile": EXPORT_FILE.name,
            "sampleFile": SAMPLE_FILE.name,
            "routingFile": ROUTING_FILE.name,
            "progressFile": PROGRESS_FILE.name,
            "defaultLine": default_line,
            "defaultDate": default_date,
            "routing": routing,
            "speedByLine": speed_by_line,
            "speedByModel": {
                "|".join(key): value for key, value in speed_by_model.items()
            },
            "progressMat": prog_mat,
            "progressLot": prog_lot,
            "ct": load_ct(),
            "rows": rows,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    template = TEMPLATE_FILE.read_text(encoding="utf-8")
    return (
        template
        .replace("%%EXPORT_FILE_NAME%%", escape(EXPORT_FILE.name))
        .replace("%%PAYLOAD%%", payload)
    )


def main():
    rows = build_rows()
    OUTPUT_FILE.write_text(render_html(rows), encoding="utf-8")
    print(f"Wrote {OUTPUT_FILE} with {len(rows)} rows")


if __name__ == "__main__":
    main()
