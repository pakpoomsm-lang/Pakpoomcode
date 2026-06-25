"""
Expander First Lot Check — Standalone Server port 3001
"""

import json, os, re, sqlite3
from datetime import datetime, timedelta
from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse, FileResponse
import uvicorn

# ── Paths ──────────────────────────────────────────────────────────────────────
_FL_DIR     = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Server_firstlot\public"
_PC_DIR     = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\JPEG FILE"
_ITEMS_JSON = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Server_firstlot\public\data\items.json"
EXPANDER_DB  = r"W:\PD\2.HEAT INDOOR\12.Pakpoom\mecp-python\expander_records.db"
MASTER_XLSX  = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Master data HEI.xlsx"
PEOPLE_XLSX  = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Data Man MC Mat.xlsx"
_CUT_DATA_DIR = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\data"

# ── Master Data ────────────────────────────────────────────────────────────────
def build_items_json():
    try:
        import openpyxl
    except ImportError:
        if os.path.exists(_ITEMS_JSON):
            with open(_ITEMS_JSON, encoding='utf-8') as f:
                return len(json.load(f))
        return 0

    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    SHEETS = ['Master UPD', 'Master Review-2025', 'Master Review-2025 (2)']
    all_rows = []
    for sh in SHEETS:
        if sh in wb.sheetnames:
            all_rows.extend(list(wb[sh].iter_rows(values_only=True))[3:])
    wb.close()

    items = {}
    for row in all_rows:
        if not row or len(row) < 6: continue
        item_code = row[5]
        if not item_code: continue
        item_code = str(item_code).strip().upper()
        if not item_code or item_code == 'ITEM': continue

        def _cell(col, default=''):
            if len(row) <= col or row[col] is None: return default
            return row[col]

        pipe_raw = str(_cell(9, '')).strip()
        pipe_m   = re.match(r'(\d+)', pipe_raw)
        pipe     = pipe_m.group(1) if pipe_m else '5'
        try:   el = float(_cell(12)) if _cell(12) != '' else None
        except: el = None
        expand_mc    = str(_cell(13, '')).strip()
        fin_press_mc = next((str(row[c]).strip() for c in [18,19,20] if len(row)>c and row[c]), '')
        hairpin_item = str(_cell(21, '')).strip()
        fin_item     = str(_cell(49, '')).strip()
        fin_desc     = str(_cell(50, '')).strip()
        try:   pitch = float(_cell(51)) if _cell(51) != '' else ''
        except: pitch = ''

        def _sp(val):
            v = str(val or '').strip()
            if not v or v.upper() == 'NA': return 'NA'
            return v if v.upper().endswith('-F') else v + '-F'

        sp1 = _sp(_cell(94, ''))
        sp2 = _sp(_cell(95, ''))
        sp3 = _sp(_cell(96, ''))
        spec_a, spec_b = ('14~15', '12~14') if pipe == '7' else ('11~12', '9~10')

        def _cell_str(col):
            if len(row) <= col or row[col] is None: return ''
            return str(row[col]).strip()

        items[item_code] = {
            'type': str(row[6] or '').strip(), 'model': str(row[7] or '').strip(),
            'pipe': pipe, 'desc': str(row[10] or '').strip(), 'el': el,
            'expandMC': expand_mc, 'finPressMC': fin_press_mc,
            'hairpinItem': hairpin_item, 'finItem': fin_item, 'finDesc': fin_desc,
            'pitch': pitch, 'specA': spec_a, 'specB': spec_b,
            'specC': '<1.5', 'specD': '<20%',
            'sp1Item': sp1, 'sp2Item': sp2, 'sp3Item': sp3,
            # ── Cutting-specific fields ─────────────────────────────────
            'cuttingPcs':    _cell_str(218),
            'laydownSide':   _cell_str(219),
            'laydownMiddle': _cell_str(220),
            'cuttingMC':     _cell_str(221),
            'insulQty':      _cell_str(225),
            'insulItem':     _cell_str(226),   # col HS
            'insulItem2':    _cell_str(227),   # col HT — second insulation item
        }

    with open(_ITEMS_JSON, 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

    # ── Also save cutting master to cutting data dir ───────────────────
    os.makedirs(_CUT_DATA_DIR, exist_ok=True)
    _cut_master = os.path.join(_CUT_DATA_DIR, 'master_items.json')
    with open(_cut_master, 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

    return len(items)

def build_cutting_people():
    """Read Data Man MC Mat.xlsx → machines.json, employees.json, checkers.json"""
    try:
        import openpyxl as _xl
    except ImportError:
        print("[People] WARNING: openpyxl not available, skipping people rebuild")
        return

    if not os.path.exists(PEOPLE_XLSX):
        print(f"[People] WARNING: ไม่พบไฟล์ {PEOPLE_XLSX}")
        return

    wb = _xl.load_workbook(PEOPLE_XLSX, read_only=True, data_only=True)
    os.makedirs(_CUT_DATA_DIR, exist_ok=True)

    # ── Sheet: Machine ─────────────────────────────────────────────────
    machines = {}
    if 'Machine' in wb.sheetnames:
        rows = list(wb['Machine'].iter_rows(values_only=True))
        for row in rows[1:]:   # row 0 = header
            if not row or row[0] is None: continue
            code = str(row[0]).strip()
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if code:
                machines[code] = name
    with open(os.path.join(_CUT_DATA_DIR, 'machines.json'), 'w', encoding='utf-8') as f:
        json.dump(machines, f, ensure_ascii=False, indent=2)

    # ── Sheet: Name Checker → employees ───────────────────────────────
    employees = {}
    checker_sheet = next((s for s in wb.sheetnames if 'checker' in s.lower()), None)
    if checker_sheet:
        rows = list(wb[checker_sheet].iter_rows(values_only=True))
        for row in rows[2:]:   # row 0 = title, row 1 = headers
            if not row or len(row) < 3: continue
            emp_id = str(row[1]).strip() if row[1] is not None else ''
            name   = str(row[2]).strip() if row[2] is not None else ''
            if emp_id and emp_id.isdigit():
                employees[emp_id] = name
    with open(os.path.join(_CUT_DATA_DIR, 'employees.json'), 'w', encoding='utf-8') as f:
        json.dump(employees, f, ensure_ascii=False, indent=2)

    # ── Sheet: Name confirm → checkers ────────────────────────────────
    checkers = {}
    confirm_sheet = next((s for s in wb.sheetnames if 'confirm' in s.lower()), None)
    if confirm_sheet:
        rows = list(wb[confirm_sheet].iter_rows(values_only=True))
        for row in rows[2:]:   # row 0 = title, row 1 = headers
            if not row or len(row) < 3: continue
            chk_id = str(row[1]).strip() if row[1] is not None else ''
            raw_name = str(row[2]).strip() if row[2] is not None else ''
            role   = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            # Strip role suffix in parentheses from name if present
            import re as _re
            name_clean = _re.sub(r'\s*\(.*?\)\s*$', '', raw_name).strip()
            if chk_id:
                checkers[chk_id] = {'name': name_clean or raw_name, 'role': role}
    with open(os.path.join(_CUT_DATA_DIR, 'checkers.json'), 'w', encoding='utf-8') as f:
        json.dump(checkers, f, ensure_ascii=False, indent=2)

    wb.close()
    print(f"[People] machines={len(machines)}, employees={len(employees)}, checkers={len(checkers)}")


try:
    n = build_items_json()
    print(f"[Master] โหลด {n} items สำเร็จ")
except Exception as e:
    print(f"[Master] WARNING: {e}")

try:
    build_cutting_people()
except Exception as e:
    print(f"[People] WARNING: {e}")

# ── Database ───────────────────────────────────────────────────────────────────
def get_db():
    con = sqlite3.connect(EXPANDER_DB)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    con = sqlite3.connect(EXPANDER_DB)
    con.execute('''CREATE TABLE IF NOT EXISTS records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT, time TEXT,
        checker TEXT, checker_id TEXT, confirm_checker TEXT,
        machine TEXT, line TEXT, seq TEXT, qty TEXT,
        item_fg TEXT, item_type TEXT,
        sp1 TEXT, sp2 TEXT, sp3 TEXT,
        val_el TEXT, val_a TEXT, val_b TEXT, val_c TEXT, val_d TEXT,
        flat_pct TEXT, ap_hairpin TEXT, ap_fin TEXT,
        pitch TEXT, hairpin_item TEXT, fin_item TEXT,
        mark_color TEXT, mark_status TEXT, remark TEXT, status TEXT,
        saved_at TEXT
    )''')
    con.commit()
    existing = {r[1] for r in con.execute("PRAGMA table_info(records)").fetchall()}
    for col in ['flat_pct', 'mark_status']:
        if col not in existing:
            con.execute(f'ALTER TABLE records ADD COLUMN {col} TEXT')
    con.commit()
    con.close()

init_db()

# ── App ────────────────────────────────────────────────────────────────────────
app = FastAPI()

# ── Pages ──────────────────────────────────────────────────────────────────────
@app.get("/")
@app.get("/expander/firstlot")
@app.get("/expander/firstlot/")
async def page_index():
    return FileResponse(os.path.join(_FL_DIR, "index.html"))

@app.get("/Layout_HEI")
@app.get("/Layout_HEI/")
async def page_layout_hei():
    return FileResponse(
        os.path.join(_FL_DIR, "layout_hei.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"}
    )

@app.get("/Layout_HEI/images/{filename:path}")
async def layout_hei_images(filename: str):
    path = os.path.join(_FL_DIR, "images", filename)
    return FileResponse(path) if os.path.exists(path) else JSONResponse({"error": "not found"}, status_code=404)

# ── Layout HEI — server-side position storage (shared across all PCs) ──────────
_LAYOUT_POS_FILE = os.path.join(_FL_DIR, "data", "layout_positions.json")

@app.get("/Layout_HEI/positions")
async def layout_get_positions():
    """Return saved machine-label positions (shared across all clients)."""
    try:
        if os.path.exists(_LAYOUT_POS_FILE):
            with open(_LAYOUT_POS_FILE, 'r', encoding='utf-8') as f:
                return JSONResponse(json.loads(f.read()))
        return JSONResponse({})
    except Exception as e:
        return JSONResponse({}, status_code=200)

@app.post("/Layout_HEI/positions")
async def layout_save_positions(request: Request):
    """Save machine-label positions to server (persists for all clients)."""
    try:
        body = await request.json()
        os.makedirs(os.path.dirname(_LAYOUT_POS_FILE), exist_ok=True)
        with open(_LAYOUT_POS_FILE, 'w', encoding='utf-8') as f:
            f.write(json.dumps(body, ensure_ascii=False, indent=2))
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.delete("/Layout_HEI/positions")
async def layout_reset_positions():
    """Delete saved positions — all clients revert to HTML defaults."""
    try:
        if os.path.exists(_LAYOUT_POS_FILE):
            os.remove(_LAYOUT_POS_FILE)
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/expander/firstlot/dashboard")
@app.get("/dashboard/firstlot")
async def page_dashboard():
    return FileResponse(os.path.join(_FL_DIR, "dashboard.html"))

@app.get("/expander/firstlot/admin")
@app.get("/firstlot/admin")
async def page_admin():
    return FileResponse(os.path.join(_FL_DIR, "admin.html"))

# ── Static ─────────────────────────────────────────────────────────────────────
@app.get("/expander/firstlot/data/{filename:path}")
async def static_data(filename: str):
    path = os.path.join(_FL_DIR, "data", filename)
    return FileResponse(path) if os.path.exists(path) else JSONResponse({"error": "not found"}, status_code=404)

@app.get("/expander/firstlot/images/{filename:path}")
async def static_images(filename: str):
    path = os.path.join(_FL_DIR, "images", filename)
    return FileResponse(path) if os.path.exists(path) else JSONResponse({"error": "not found"}, status_code=404)

@app.get("/expander/firstlot/process-control/{filename:path}")
async def static_pc(filename: str):
    for fn in [filename, filename.replace('.jpg', '.JPG'), filename.replace('.JPG', '.jpg')]:
        path = os.path.join(_PC_DIR, fn)
        if os.path.exists(path):
            return FileResponse(path)
    return JSONResponse({"error": "not found", "file": filename}, status_code=404)

# ── API ────────────────────────────────────────────────────────────────────────
@app.get("/expander/firstlot/api/reload-master")
async def api_reload():
    try:
        n = build_items_json()
        return JSONResponse({'ok': True, 'items': n, 'message': f'โหลด Master สำเร็จ {n} items'})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/expander/firstlot/api/records")
async def api_records():
    con = get_db()
    rows = [dict(r) for r in con.execute("SELECT * FROM records ORDER BY id DESC LIMIT 500").fetchall()]
    con.close()
    return JSONResponse(rows)

@app.get("/expander/firstlot/api/records/all")
async def api_records_all():
    con = get_db()
    rows = [dict(r) for r in con.execute("SELECT * FROM records ORDER BY id DESC").fetchall()]
    con.close()
    return JSONResponse(rows)

@app.post("/expander/firstlot/api/record")
async def api_save(request: Request):
    d = await request.json()
    try:
        con = get_db()
        con.execute('''
            INSERT INTO records
            (date,time,checker,checker_id,confirm_checker,
             machine,line,seq,qty,item_fg,item_type,
             sp1,sp2,sp3,val_el,val_a,val_b,val_c,val_d,flat_pct,
             ap_hairpin,ap_fin,pitch,hairpin_item,fin_item,
             mark_color,mark_status,remark,status,saved_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            d.get('date'),      d.get('time'),
            d.get('checker'),   d.get('checkerID'),   d.get('confirmChecker'),
            d.get('machine'),   d.get('line'),        d.get('seq'),   d.get('qty'),
            d.get('itemFG'),    d.get('type'),
            d.get('sp1'),       d.get('sp2'),         d.get('sp3'),
            d.get('valEL'),     d.get('valA'),        d.get('valB'),
            d.get('valC'),      d.get('valD'),        d.get('flatPct'),
            d.get('apHairpin'), d.get('apFIN'),
            d.get('pitch'),     d.get('hairpin'),     d.get('fin'),
            d.get('markColor'), d.get('markStatus'),
            d.get('remark'),    d.get('status'),
            datetime.now().isoformat()
        ))
        con.commit()
        rid = con.execute('SELECT last_insert_rowid()').fetchone()[0]
        con.close()
        return JSONResponse({'ok': True, 'id': rid})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.put("/expander/firstlot/api/record/{rid}")
async def api_update(rid: int, request: Request):
    d = await request.json()
    allowed = [
        'date','time','checker','checker_id','confirm_checker',
        'machine','line','seq','qty','item_fg','item_type',
        'sp1','sp2','sp3','val_el','val_a','val_b','val_c','val_d','flat_pct',
        'ap_hairpin','ap_fin','pitch','hairpin_item','fin_item',
        'mark_color','mark_status','status','remark'
    ]
    sets = ', '.join(f'{c}=?' for c in allowed)
    vals = [d.get(c, '') for c in allowed] + [rid]
    try:
        con = get_db()
        con.execute(f'UPDATE records SET {sets} WHERE id=?', vals)
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.delete("/expander/firstlot/api/record/{rid}")
async def api_delete(rid: int):
    try:
        con = get_db()
        con.execute('DELETE FROM records WHERE id=?', (rid,))
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.post("/expander/firstlot/api/records/bulk-delete")
async def api_bulk_delete(request: Request):
    try:
        ids = (await request.json()).get('ids', [])
        if not ids: return JSONResponse({'ok': False, 'error': 'No IDs provided'})
        con = get_db()
        placeholders = ','.join('?' * len(ids))
        con.execute(f'DELETE FROM records WHERE id IN ({placeholders})', ids)
        con.commit(); con.close()
        return JSONResponse({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/expander/firstlot/api/export")
async def api_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        con = get_db()
        rows = [dict(r) for r in con.execute('SELECT * FROM records ORDER BY id').fetchall()]
        con.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'First Lot Check'

        def avg_from_str(s):
            try:
                nums = [float(x) for x in str(s or '').split(',') if x.strip()]
                return round(sum(nums)/len(nums), 2) if nums else ''
            except: return ''

        headers = ['No','Date','Time','Checker','Confirm','Machine','Line','Seq','Qty',
                   'Item FG','Type','SP1','SP2','SP3',
                   'EL','EL Avg','A','A Avg','B','B Avg',
                   'C','D','D%','Hairpin','FIN','Pitch','Hairpin Item','FIN Item',
                   'Mark Color','Mark Status','Remark','Status']
        ws.append(headers)
        thin = Side(style='thin', color='B0BEC5')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='0D47A1')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center')
        pf = PatternFill('solid', fgColor='C8E6C9')
        nf = PatternFill('solid', fgColor='FFCDD2')
        for i, r in enumerate(rows, 1):
            ws.append([i, r.get('date'), r.get('time'),
                r.get('checker'), r.get('confirm_checker'), r.get('machine'),
                r.get('line'), r.get('seq'), r.get('qty'),
                r.get('item_fg'), r.get('item_type'),
                r.get('sp1'), r.get('sp2'), r.get('sp3'),
                r.get('val_el'), avg_from_str(r.get('val_el')),
                r.get('val_a'),  avg_from_str(r.get('val_a')),
                r.get('val_b'),  avg_from_str(r.get('val_b')),
                r.get('val_c'), r.get('val_d'), r.get('flat_pct'),
                r.get('ap_hairpin'), r.get('ap_fin'),
                r.get('pitch'), r.get('hairpin_item'), r.get('fin_item'),
                r.get('mark_color'), r.get('mark_status'),
                r.get('remark'), r.get('status')])
            for cell in ws[i+1]: cell.border = bdr
            st = str(r.get('status') or '')
            ws.cell(i+1, len(headers)).fill = (
                pf if 'PASS' in st else (nf if 'FAIL' in st else PatternFill()))
        for idx, w in enumerate(
            [5,10,8,18,18,22,6,6,8,14,8,14,14,14,10,8,10,8,10,8,10,12,8,8,14,16,16,12,16,20,10], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = ws.dimensions
        out = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\First lot machine\Expander\expander_export.xlsx"
        os.makedirs(os.path.dirname(out), exist_ok=True)
        wb.save(out)
        return JSONResponse({'ok': True, 'file': out, 'rows': len(rows)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.delete("/expander/firstlot/api/records/delete-all")
async def api_delete_all(request: Request):
    try:
        body = await request.json()
        if body.get('confirm') != 'DELETE_ALL':
            return JSONResponse({'ok': False, 'error': 'confirm required'}, status_code=400)
        con = get_db()
        con.execute('DELETE FROM records')
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

# ── Cutting Check ──────────────────────────────────────────────────────────────
_CUT_DIR = r"W:\PD\2.HEAT INDOOR\13.Suphamat P"
_OLD_CUT_DB = os.path.join(_CUT_DIR, "expander_records.db")   # legacy shared location
_CUT_DB  = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Server_firstlot\cutting_records.db"
_JPEG_DIR = os.path.join(_CUT_DIR, "JPEG FILE")

def _sync_table_from_old(new_con, table, old_db_path):
    """Copy any rows from the old DB that aren't in this new DB yet (matched by id).
    Gap-free + idempotent: safe to run now and again at every restart. IDs are
    preserved so AUTOINCREMENT keeps counting up. The old DB is never modified."""
    try:
        if not os.path.exists(old_db_path):
            return
        old = sqlite3.connect(old_db_path)
        if not old.execute("SELECT name FROM sqlite_master "
                           "WHERE type='table' AND name=?", (table,)).fetchone():
            old.close(); return
        new_max  = new_con.execute(f"SELECT COALESCE(MAX(id),0) FROM {table}").fetchone()[0]
        new_cols = [r[1] for r in new_con.execute(f"PRAGMA table_info({table})").fetchall()]
        cur = old.execute(f"SELECT * FROM {table} WHERE id > ?", (new_max,))
        old_cols = [d[0] for d in cur.description]
        use = [c for c in old_cols if c in new_cols]          # only shared columns
        idx = [old_cols.index(c) for c in use]
        rows = [tuple(r[i] for i in idx) for r in cur.fetchall()]
        old.close()
        if not rows:
            return
        ph = ','.join(['?'] * len(use))
        new_con.executemany(
            f"INSERT INTO {table} ({','.join(use)}) VALUES ({ph})", rows)
        new_con.commit()
        print(f"[Migrate] {table}: synced {len(rows)} new rows from {old_db_path}")
    except Exception as e:
        print(f"[Migrate] WARNING {table}: {e}")

def cutting_get_db():
    con = sqlite3.connect(_CUT_DB)
    con.row_factory = sqlite3.Row
    return con

def cutting_init_db():
    con = sqlite3.connect(_CUT_DB)
    con.execute('''CREATE TABLE IF NOT EXISTS cutting_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_card TEXT, mc_line TEXT,
        tag_fg_no1 TEXT, tag_fg_no2 TEXT,
        item_insulation TEXT, item_type TEXT,
        line TEXT, seq TEXT, lot TEXT,
        check_laydown TEXT, check_burr TEXT, check_drawing TEXT,
        check_copper TEXT, check_apprence TEXT,
        insulation_qty TEXT, cutting_pcs TEXT,
        lay_down_side INTEGER DEFAULT 0,
        lay_down_middle INTEGER DEFAULT 0,
        leader_confirm TEXT,
        status TEXT,
        saved_at TEXT
    )''')
    con.commit()
    # Migrate existing table — add new columns if missing
    existing = {r[1] for r in con.execute("PRAGMA table_info(cutting_records)").fetchall()}
    for col in ['vendor_insulation', 'tag_insulation', 'item_insulation_2']:
        if col not in existing:
            con.execute(f'ALTER TABLE cutting_records ADD COLUMN {col} TEXT')
    con.commit()
    # Pull any rows still only in the old shared DB into this Server_firstlot DB
    _sync_table_from_old(con, 'cutting_records', _OLD_CUT_DB)
    con.close()

cutting_init_db()

# ── Cutting Pages ──────────────────────────────────────────────────────────────
@app.get("/cutting")
@app.get("/cutting/")
@app.get("/cutting/firstlot")
@app.get("/cutting/firstlot/")
async def page_cutting():
    return FileResponse(os.path.join(_CUT_DIR, "cutting.html"))

# ── Cutting Static ─────────────────────────────────────────────────────────────
@app.get("/cutting/data/{filename:path}")
async def cutting_static_data(filename: str):
    path = os.path.join(_CUT_DIR, "data", filename)
    if os.path.exists(path):
        return FileResponse(path, headers={"Cache-Control": "no-store, no-cache, must-revalidate"})
    return JSONResponse({"error": "not found"}, status_code=404)

@app.get("/images/{filename:path}")
async def cutting_static_images(filename: str):
    path = os.path.join(_CUT_DIR, "images", filename)
    return FileResponse(path) if os.path.exists(path) else JSONResponse({"error": "not found"}, status_code=404)

@app.get("/cutting/process-control/{filename:path}")
async def cutting_process_control(filename: str):
    for fn in [filename, filename.replace('.jpg', '.JPG'), filename.replace('.JPG', '.jpg')]:
        path = os.path.join(_JPEG_DIR, fn)
        if os.path.exists(path):
            return FileResponse(path)
    return JSONResponse({"error": "not found", "file": filename}, status_code=404)

# ── Cutting API ────────────────────────────────────────────────────────────────
@app.get("/cutting/api/reload-master")
async def cutting_api_reload_master():
    try:
        n = build_items_json()
        build_cutting_people()
        return JSONResponse({'ok': True, 'items': n, 'message': f'โหลด Master + People สำเร็จ {n} items'})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/cutting/api/cutting-records")
async def cutting_api_records():
    con = cutting_get_db()
    rows = [dict(r) for r in con.execute(
        "SELECT * FROM cutting_records ORDER BY id DESC").fetchall()]
    con.close()
    return JSONResponse(rows)

@app.post("/cutting/api/cutting-record")
async def cutting_api_save(request: Request):
    d = await request.json()
    try:
        con = cutting_get_db()
        con.execute('''INSERT INTO cutting_records
            (id_card, mc_line, tag_fg_no1, tag_fg_no2, item_insulation,
             item_type, line, seq, lot,
             check_laydown, check_burr, check_drawing, check_copper, check_apprence,
             insulation_qty, cutting_pcs, lay_down_side, lay_down_middle,
             leader_confirm, status, saved_at,
             vendor_insulation, tag_insulation, item_insulation_2)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            d.get('idCard'),        d.get('mcLine'),
            d.get('tagFgNo1'),      d.get('tagFgNo2'),      d.get('itemInsulation'),
            # 'itemType' is sent by the frontend (camelCase); 'type' is legacy fallback
            d.get('itemType') or d.get('type', ''),
            d.get('line'),          d.get('seq'),    d.get('lot'),
            d.get('checkLaydown'),  d.get('checkBurr'),     d.get('checkDrawing'),
            d.get('checkCopper'),   d.get('checkApprence'),
            d.get('insulationQty'), d.get('cuttingPcs'),
            '1' if d.get('layDownSide')   else '0',
            '1' if d.get('layDownMiddle') else '0',
            d.get('leaderConfirm'), d.get('status'),
            datetime.now().isoformat(),
            d.get('vendorInsulation', ''), d.get('tagInsulation', ''),
            d.get('itemInsulation2', '')
        ))
        con.commit()
        rid = con.execute('SELECT last_insert_rowid()').fetchone()[0]
        con.close()
        return JSONResponse({'ok': True, 'id': rid})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.put("/cutting/api/cutting-record/{rid}")
async def cutting_api_update(rid: int, request: Request):
    d = await request.json()
    allowed = ['id_card', 'mc_line', 'tag_fg_no1', 'tag_fg_no2',
               'item_insulation', 'item_insulation_2',
               'vendor_insulation', 'tag_insulation',
               'item_type', 'line', 'seq', 'lot', 'leader_confirm', 'status']
    sets = ', '.join(f'{c}=?' for c in allowed)
    vals = [d.get(c, '') for c in allowed] + [rid]
    try:
        con = cutting_get_db()
        con.execute(f'UPDATE cutting_records SET {sets} WHERE id=?', vals)
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.delete("/cutting/api/cutting-record/{rid}")
async def cutting_api_delete(rid: int):
    try:
        con = cutting_get_db()
        con.execute('DELETE FROM cutting_records WHERE id=?', (rid,))
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.post("/cutting/api/cutting-records/bulk-delete")
async def cutting_api_bulk_delete(request: Request):
    try:
        ids = (await request.json()).get('ids', [])
        if not ids: return JSONResponse({'ok': False, 'error': 'No IDs provided'})
        con = cutting_get_db()
        placeholders = ','.join('?' * len(ids))
        con.execute(f'DELETE FROM cutting_records WHERE id IN ({placeholders})', ids)
        con.commit(); con.close()
        return JSONResponse({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/cutting/api/cutting-export-excel")
async def cutting_api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from fastapi.responses import StreamingResponse
        import io
        con = cutting_get_db()
        rows = [dict(r) for r in con.execute(
            'SELECT * FROM cutting_records ORDER BY id').fetchall()]
        con.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cutting Records'
        headers = [
            'No', 'Saved At', 'ID Card', 'MC/Line', 'Type',
            'Tag FG No.1', 'Tag FG No.2',
            'Item Insulation 1', 'Item Insulation 2', 'Vendor Insulation',
            'Line', 'Seq', 'Lot',
            'Lay Down', 'Burr', 'Drawing', 'Copper', 'Apprence',
            'Insul.Qty', 'Cut.Pcs',
            'Lay Side', 'Lay Mid', 'Leader Confirm', 'Status'
        ]
        ws.append(headers)
        thin = Side(style='thin', color='B0BEC5')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='1565C0')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center')
        pf = PatternFill('solid', fgColor='C8E6C9')
        nf = PatternFill('solid', fgColor='FFCDD2')
        STATUS_COL = len(headers)
        for i, r in enumerate(rows, 1):
            ws.append([
                i,
                r.get('saved_at'),           r.get('id_card'),          r.get('mc_line'),
                r.get('item_type', ''),
                r.get('tag_fg_no1'),         r.get('tag_fg_no2'),
                r.get('item_insulation'),    r.get('item_insulation_2', ''),
                r.get('vendor_insulation', ''),
                r.get('line'),               r.get('seq'),              r.get('lot'),
                r.get('check_laydown'),      r.get('check_burr'),       r.get('check_drawing'),
                r.get('check_copper'),       r.get('check_apprence'),
                r.get('insulation_qty'),     r.get('cutting_pcs'),
                '✓' if r.get('lay_down_side')   else '',
                '✓' if r.get('lay_down_middle') else '',
                r.get('leader_confirm'),     r.get('status')
            ])
            for cell in ws[i+1]: cell.border = bdr
            st = str(r.get('status') or '')
            ws.cell(i+1, STATUS_COL).fill = pf if st == 'OK' else (nf if st == 'NG' else PatternFill())
        col_widths = [5, 20, 16, 14, 10, 16, 16, 16, 16, 14, 6, 6, 8, 8, 8, 10, 8, 10, 9, 9, 9, 9, 18, 8]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = ws.dimensions
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="cutting_records.xlsx"'})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

# ── HP Check ──────────────────────────────────────────────────────────────────
_OLD_HP_DB = EXPANDER_DB   # legacy: hp_records lived in the shared Expander DB
_HP_DB     = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Server_firstlot\hp_records.db"

def hp_get_db():
    con = sqlite3.connect(_HP_DB)
    con.row_factory = sqlite3.Row
    return con

def hp_init_db():
    con = sqlite3.connect(_HP_DB)
    con.execute('''CREATE TABLE IF NOT EXISTS hp_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts TEXT, save_type TEXT,
        line TEXT, seq TEXT, prod_month TEXT, lot TEXT,
        item TEXT, item_fg TEXT, item_type TEXT, cu_item TEXT, die TEXT,
        mc TEXT, checker TEXT,
        coils INTEGER, coil_data TEXT, coil_size TEXT,
        avg_l1 REAL, avg_l2 REAL, avg_diff REAL, avg_flat REAL,
        vendor TEXT,
        meas_full TEXT, measurements TEXT, visuals TEXT,
        result TEXT, confirm_by TEXT,
        saved_at TEXT
    )''')
    con.commit()
    # Pull any rows still only in the old shared Expander DB into this Server_firstlot DB
    _sync_table_from_old(con, 'hp_records', _OLD_HP_DB)
    con.close()

hp_init_db()

@app.get("/hp/firstlot")
@app.get("/hp/firstlot/")
async def page_hp():
    return FileResponse(os.path.join(_FL_DIR, "hp_checker.html"))

@app.get("/hp/api/records")
async def hp_api_records():
    con = hp_get_db()
    rows = [dict(r) for r in con.execute(
        "SELECT * FROM hp_records ORDER BY id DESC").fetchall()]
    con.close()
    return JSONResponse(rows)

@app.get("/hp/api/export")
async def hp_api_export():
    try:
        import openpyxl, json as _json
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        con = hp_get_db()
        rows = [dict(r) for r in con.execute(
            'SELECT * FROM hp_records ORDER BY id ASC').fetchall()]
        con.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'HP Check'

        thin = Side(style='thin', color='B0BEC5')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── 99 columns ────────────────────────────────────────────────
        # A-L   (12): Base info
        # M-BH  (42): Coil No.1–14 + Sale ID + Black point (3 cols × 14)
        # BI-CD (42): Line 1–14 measurements (3 cols × 14)
        # CE-CG  (3): Appearance checks

        # Row 1: main headers
        h1 = ['Date','Time','Line','Seq','Prod/month',
               'Machine Hairpin Bender',
               'Item Hairpin','Item FG','Type','Item Copper',
               'Coil Size','Vender']
        for n in range(1, 15):
            h1 += [f'Coil No. {n}', 'Sale ID', 'Black point']
        for n in range(1, 15):
            h1 += [f'Line {n}', None, 'Flatness']
        h1 += ['Apprence check', None, None]
        ws.append(h1)

        # Row 2: sub-headers
        h2 = [None] * 54   # 12 base + 42 coil/saleid/bp = 54, all merged vertically
        for _ in range(14):
            h2 += ['L1 ±1 mm.\nL2 ±1 mm.', 'D\n0-1 mm.', '≤20 %\nOK.']
        h2 += ['ไม่มีรอยขีดบริเวณโค้งท่อ',
               'รอยย่นไม่เกิน limit กำหนด',
               'ผู้ลงบันทึก']
        ws.append(h2)

        # ── Merge cells ───────────────────────────────────────────────
        # Cols 1-54: base + coil columns merge rows 1-2
        for col in range(1, 55):
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=2,   end_column=col)
        # Each "Line N": merge 2 cols in row 1 (L1/L2 + D); Flatness separate
        for n in range(14):
            sc = 55 + n * 3
            ws.merge_cells(start_row=1, start_column=sc,
                           end_row=1,   end_column=sc + 1)
        # "Apprence check": merge 3 cols in row 1
        ws.merge_cells(start_row=1, start_column=97,
                       end_row=1,   end_column=99)

        # ── Style headers ─────────────────────────────────────────────
        hf1 = PatternFill('solid', fgColor='1B5E20')
        hf2 = PatternFill('solid', fgColor='2E7D32')
        fw  = Font(bold=True, color='FFFFFF', size=9)
        for row_idx, hfill in ((1, hf1), (2, hf2)):
            for col in range(1, 100):
                c = ws.cell(row_idx, col)
                c.fill = hfill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center',
                                        wrap_text=True)
                if c.value:
                    c.font = fw
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 34

        def mc_name(mc):
            if mc and mc.upper().startswith('HB'):
                return 'Hairpin Bender No.' + mc[2:]
            return mc or ''

        # ── Data rows (one row per inspection record) ─────────────────
        pf = PatternFill('solid', fgColor='C8E6C9')
        nf = PatternFill('solid', fgColor='FFCDD2')

        for r in rows:
            try: meas_full = _json.loads(r.get('meas_full') or '[]')
            except: meas_full = []
            try: coil_data = _json.loads(r.get('coil_data') or '[]')
            except: coil_data = []
            try: visuals   = _json.loads(r.get('visuals')   or '[]')
            except: visuals = []

            ts = str(r.get('ts') or '')
            date_str = ts[:10] if ts else ''
            time_str = ts[11:19] if len(ts) > 10 else ''

            row_data = [
                date_str, time_str,
                r.get('line'), r.get('seq'), r.get('prod_month'),
                mc_name(r.get('mc')),
                r.get('item'), r.get('item_fg'), r.get('item_type'),
                r.get('cu_item'), r.get('coil_size'),
                r.get('vendor') or ''
            ]

            # 14 coil triplets: Coil No. n, Sale ID, Black point
            for i in range(14):
                c = coil_data[i] if i < len(coil_data) else None
                if c and isinstance(c, dict):
                    row_data += [c.get('coilNo', ''),
                                 c.get('saleId', '') or '',
                                 c.get('blackPoint', '') or '']
                else:
                    row_data += ['', '', '']

            # 14 measurement groups: L1/L2, D, Flat%
            for i in range(14):
                m = meas_full[i] if i < len(meas_full) else None
                has_coil = bool(coil_data[i]) if i < len(coil_data) else False
                if m and isinstance(m, dict) and has_coil:
                    l1 = m.get('l1'); l2 = m.get('l2')
                    l1l2 = (f"{l1}" if l1 is not None else '') + \
                           (' / ' + f"{l2}" if l2 is not None else '')
                    diff = m.get('diff')
                    flat = m.get('flat')
                    row_data += [l1l2.strip(' /'), diff,
                                 f"{flat}%" if flat is not None else '']
                else:
                    row_data += ['', '', '']

            # Appearance checks
            vis_list = [v for v in visuals if v in ('OK', 'NG')]
            has_ng   = any(v == 'NG' for v in vis_list)
            vis_str  = 'NG' if has_ng else ('OK' if vis_list else '')
            row_data += [vis_str, vis_str, r.get('checker') or '']

            ws.append(row_data)
            ri = ws.max_row
            for col in range(1, 100):
                c = ws.cell(ri, col)
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center',
                                        wrap_text=False)
            app_fill = (pf if vis_str == 'OK' else
                        nf if vis_str == 'NG' else PatternFill())
            ws.cell(ri, 97).fill = app_fill
            ws.cell(ri, 98).fill = app_fill

        # ── Column widths ─────────────────────────────────────────────
        base_w = [10, 9, 5, 5, 10, 22, 20, 20, 7, 16, 9, 10]
        coil_w = [16, 18, 7] * 14
        meas_w = [12, 6, 7] * 14
        app_w  = [20, 20, 14]
        for idx, w in enumerate(base_w + coil_w + meas_w + app_w, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        ws.freeze_panes = 'C3'
        ws.auto_filter.ref = f"A2:{get_column_letter(99)}{max(ws.max_row, 2)}"

        out = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\First lot machine\Hairpin Bender\hp_export.xlsx"
        os.makedirs(os.path.dirname(out), exist_ok=True)
        wb.save(out)
        return JSONResponse({'ok': True, 'file': out, 'rows': len(rows)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.post("/hp/api/record")
async def hp_api_save(request: Request):
    d = await request.json()
    try:
        con = hp_get_db()
        con.execute('''INSERT INTO hp_records
            (ts, save_type, line, seq, prod_month, lot,
             item, item_fg, item_type, cu_item, die,
             mc, checker, coils, coil_data, coil_size,
             avg_l1, avg_l2, avg_diff, avg_flat, vendor,
             meas_full, measurements, visuals, result, confirm_by, saved_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            d.get('ts'), d.get('saveType'),
            d.get('line'), d.get('seq'), d.get('prodMonth'), d.get('lot'),
            d.get('item'), d.get('itemFG'), d.get('type'), d.get('cuItem'), d.get('die'),
            d.get('mc'), d.get('checker'),
            d.get('coils'),
            json.dumps(d.get('coilData') or [], ensure_ascii=False),
            d.get('coilSize'),
            d.get('avgL1'), d.get('avgL2'), d.get('avgDiff'), d.get('avgFlat'),
            d.get('vendor'),
            json.dumps(d.get('measFull') or [], ensure_ascii=False),
            json.dumps(d.get('measurements') or [], ensure_ascii=False),
            json.dumps(d.get('visuals') or [], ensure_ascii=False),
            d.get('result'), d.get('confirmBy'),
            datetime.now().isoformat()
        ))
        con.commit()
        rid = con.execute('SELECT last_insert_rowid()').fetchone()[0]
        con.close()
        return JSONResponse({'ok': True, 'id': rid})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.put("/hp/api/record/{rid}")
async def hp_api_update(rid: int, request: Request):
    d = await request.json()
    allowed = ['line','seq','mc','checker','item','item_fg','item_type',
               'cu_item','die','vendor','coil_size','result','confirm_by']
    sets = ', '.join(f'{c}=?' for c in allowed)
    vals = [d.get(c, '') for c in allowed] + [rid]
    try:
        con = hp_get_db()
        con.execute(f'UPDATE hp_records SET {sets} WHERE id=?', vals)
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.delete("/hp/api/record/{rid}")
async def hp_api_delete(rid: int):
    try:
        con = hp_get_db()
        con.execute('DELETE FROM hp_records WHERE id=?', (rid,))
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.post("/hp/api/records/bulk-delete")
async def hp_api_bulk_delete(request: Request):
    try:
        ids = (await request.json()).get('ids', [])
        if not ids: return JSONResponse({'ok': False, 'error': 'No IDs provided'})
        con = hp_get_db()
        placeholders = ','.join('?' * len(ids))
        con.execute(f'DELETE FROM hp_records WHERE id IN ({placeholders})', ids)
        con.commit(); con.close()
        return JSONResponse({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

# ══════════════════════════════════════════════════════════════════════════════
# FP FIRST LOT CONFIRMATION
# ══════════════════════════════════════════════════════════════════════════════
_FP_DIR  = _FL_DIR   # serve HTML from same public/ folder
_FP_DB   = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Server_firstlot\fp_records.db"

# ── Spec list (Master speclist.xlsm) ──────────────────────────────────────────
_FP_SPECLIST_PATH      = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Master speclist.xlsm"
_FP_SPECLIST_SHEET     = 'Speclistnew'
_FP_SPECLIST_HDR_ROW   = 2
_FP_SPECLIST_KEY_COL   = 2   # col C  — Item LV0 (FG)
_FP_SPECLIST_MAT_COL   = 6   # col G  — AL Item
_FP_SPECLIST_HOLES_COL = 11  # col L  — Fin hole
_FP_SPECLIST_DIE_COL   = 8   # col I  — Dia Pipe
_FP_SPECLIST_FP_NEG    = 15  # col P  — Fin pitch (-)
_FP_SPECLIST_FP_NOM    = 16  # col Q  — Fin pitch nominal
_FP_SPECLIST_FP_POS    = 17  # col R  — Fin pitch (+)
_FP_SPECLIST_QFIN_COL  = 13  # col N  — Q'ty Fin
_FP_SPECLIST_THICK_COL = 45  # col AT — Thickness

_fp_spec_raw      = None
_fp_item_rm_cache = None
_fp_sheet_cache   = {}   # machine / checker / confirm sheets

# ── Sheet configs (same Data Man MC Mat.xlsx) ─────────────────────────────────
_FP_SHEETS = {
    'machine': {'sheet': 'Machine',      'header_row': 0, 'name_field': 'MACHINE NAME'},
    'checker': {'sheet': 'Name Checker', 'header_row': 1, 'name_field': 'Name list'},
    'confirm': {'sheet': 'Name confirm', 'header_row': 1, 'name_field': 'Name list GL./Sub GL.3'},
}

def _fp_c(row, ci):
    try: v = row[ci]
    except IndexError: return ''
    return str(v).strip() if v is not None else ''

def _fp_load_speclist():
    global _fp_spec_raw
    if _fp_spec_raw is not None: return _fp_spec_raw, None
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(_FP_SPECLIST_PATH, read_only=True, data_only=True, keep_vba=False)
        if _FP_SPECLIST_SHEET not in wb.sheetnames:
            wb.close(); return None, f"Sheet '{_FP_SPECLIST_SHEET}' not found"
        raw = list(wb[_FP_SPECLIST_SHEET].iter_rows(values_only=True))
        wb.close()
        _fp_spec_raw = raw[_FP_SPECLIST_HDR_ROW + 1:]
        print(f"[FP] Speclist loaded — {len(_fp_spec_raw)} rows")
        return _fp_spec_raw, None
    except Exception as e: return None, str(e)

def _fp_load_item_rm():
    global _fp_item_rm_cache
    if _fp_item_rm_cache is not None: return _fp_item_rm_cache, None
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(PEOPLE_XLSX, read_only=True, data_only=True)
        if 'Item RM' not in wb.sheetnames:
            wb.close(); return None, "Sheet 'Item RM' not found"
        raw = list(wb['Item RM'].iter_rows(values_only=True))
        wb.close()
        result = {}
        for row in raw:
            key = str(row[0]).strip() if row[0] is not None else ''
            val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            if key and key not in ('None','nan'):
                result[key.upper()] = val if val not in ('None','nan') else ''
        _fp_item_rm_cache = result
        print(f"[FP] Item RM loaded — {len(result)} entries")
        return _fp_item_rm_cache, None
    except Exception as e: return None, str(e)

def _fp_find_item_rm(al_item_code):
    cache, _ = _fp_load_item_rm()
    return (cache or {}).get(al_item_code.strip().upper(), '')

def _fp_find_spec(item_fg):
    rows, err = _fp_load_speclist()
    if err: return None, err
    q = item_fg.strip().upper()
    for row in rows:
        if _fp_c(row, _FP_SPECLIST_KEY_COL).upper() == q:
            mat = _fp_c(row, _FP_SPECLIST_MAT_COL)
            mat_spec = _fp_find_item_rm(mat) if mat and mat not in ('0','None','nan') else ''
            mat_sz_r = _fp_c(row, _FP_SPECLIST_THICK_COL)
            mat_sz = mat_sz_r if mat_sz_r not in ('None','nan','0','') else mat_spec
            def cl(v): return v if v not in ('None','nan','','0') else ''
            return {
                'holes':         _fp_c(row, _FP_SPECLIST_HOLES_COL) or '',
                'fin_die':       _fp_c(row, _FP_SPECLIST_DIE_COL) or '',
                'mat_type':      mat if mat not in ('None','nan','0','') else '',
                'mat_type_spec': mat_spec,
                'mat_size':      mat_sz.strip() if mat_sz not in ('None','nan','') else '',
                'fp_nom':  cl(_fp_c(row, _FP_SPECLIST_FP_NOM)),
                'fp_neg':  cl(_fp_c(row, _FP_SPECLIST_FP_NEG)),
                'fp_pos':  cl(_fp_c(row, _FP_SPECLIST_FP_POS)),
                'q_fin':   cl(_fp_c(row, _FP_SPECLIST_QFIN_COL)),
            }, None
    return None, 'Not found'

def _fp_load_sheet(key):
    if key in _fp_sheet_cache: return _fp_sheet_cache[key], None
    cfg = _FP_SHEETS.get(key)
    if not cfg: return None, f"Unknown key: {key}"
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(PEOPLE_XLSX, read_only=True, data_only=True)
        sh = cfg['sheet']
        if sh not in wb.sheetnames:
            wb.close(); return None, f"Sheet '{sh}' not found"
        rows = list(wb[sh].iter_rows(values_only=True))
        wb.close()
        hr = cfg['header_row']
        hdrs = [str(h).strip() if h is not None else f"Col{i}" for i,h in enumerate(rows[hr])]
        data = []
        for row in rows[hr+1:]:
            if all(v is None for v in row): continue
            data.append({hdrs[i]: str(row[i]).strip() if row[i] is not None else ''
                         for i in range(min(len(hdrs), len(row)))})
        _fp_sheet_cache[key] = data
        print(f"[FP] '{sh}' loaded — {len(data)} rows")
        return data, None
    except Exception as e: return None, str(e)

def _fp_find_match(sheet_key, query):
    data, err = _fp_load_sheet(sheet_key)
    if err: return None, None, err
    q = query.strip().upper()
    nf = _FP_SHEETS[sheet_key]['name_field']
    for row in data:
        for v in row.values():
            if str(v).strip().upper() == q:
                clean = {k: v for k,v in row.items() if v and v.lower() not in ('none','nan','')}
                return clean, clean.get(nf, ''), None
    return None, None, 'Not found'

def _fp_find_hei(item_code):
    try:
        with open(_ITEMS_JSON, encoding='utf-8') as f:
            items = json.load(f)
        item = items.get(item_code.strip().upper())
        if item:
            return {'al_item': item.get('finItem', '')}, None
        return None, 'Not found'
    except Exception as e: return None, str(e)

# ── FP Database ────────────────────────────────────────────────────────────────
def _fp_init_db():
    con = sqlite3.connect(_FP_DB)
    con.execute('''CREATE TABLE IF NOT EXISTS fp_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT, time TEXT,
        mc TEXT, prod_month TEXT, lot TEXT, line TEXT, seq TEXT,
        type TEXT, item TEXT, desc1 TEXT, model TEXT,
        vendor TEXT, coil_no TEXT, pack_no TEXT, al_item TEXT, mat_type TEXT,
        appear_coil TEXT, appear_flare TEXT, appear_slit TEXT, result_appear TEXT,
        fp1 TEXT, fp2 TEXT, fp3 TEXT, fp4 TEXT, fp_avg TEXT, result_fp TEXT,
        ff1 TEXT, ff2 TEXT, ff3 TEXT, ff4 TEXT, ff_avg TEXT, result_ff TEXT,
        qty_fin TEXT, result_qty TEXT,
        fin_die TEXT, holes TEXT,
        final TEXT, checker TEXT, confirmer TEXT,
        spec_pitch TEXT, oil TEXT, spec_thick TEXT, compare_al TEXT, sc_appear TEXT,
        sc1 TEXT, sc2 TEXT, sc3 TEXT, sc4 TEXT, sc_avg TEXT, result_sc TEXT,
        order_prod TEXT, saved_at TEXT
    )''')
    con.commit()
    # Migrate existing DBs — add pack_no column if missing
    existing = {r[1] for r in con.execute("PRAGMA table_info(fp_records)").fetchall()}
    if 'pack_no' not in existing:
        con.execute('ALTER TABLE fp_records ADD COLUMN pack_no TEXT')
    con.commit(); con.close()

def _fp_get_db():
    con = sqlite3.connect(_FP_DB)
    con.row_factory = sqlite3.Row
    return con

_fp_init_db()

# preload FP caches at startup
def _fp_preload():
    import threading, time as _t
    def _load():
        _t.sleep(1)
        print("[FP] Pre-loading speclist + Item RM + lookup sheets …")
        _fp_load_speclist()
        _fp_load_item_rm()
        for k in _FP_SHEETS: _fp_load_sheet(k)
        print("[FP] Pre-load complete.")
    threading.Thread(target=_load, daemon=True).start()

_fp_preload()

_FP_IMG_DIR = os.path.join(_FL_DIR, "fp_img")

# ── FP Pages ───────────────────────────────────────────────────────────────────
@app.get("/fp/firstlot")
@app.get("/fp/firstlot/")
async def fp_page_index():
    return FileResponse(os.path.join(_FP_DIR, "fp_firstlot.html"))

@app.get("/fp/img/{filename:path}")
async def fp_img(filename: str):
    path = os.path.join(_FP_IMG_DIR, filename)
    return FileResponse(path) if os.path.exists(path) else JSONResponse({"error": "not found"}, status_code=404)

# ── FP Lookup API ──────────────────────────────────────────────────────────────
@app.get("/fp/api/ready")
async def fp_api_ready():
    return JSONResponse({'ready': _fp_spec_raw is not None})

@app.get("/fp/api/reload")
async def fp_api_reload():
    global _fp_spec_raw, _fp_item_rm_cache, _fp_sheet_cache
    _fp_spec_raw = None; _fp_item_rm_cache = None; _fp_sheet_cache = {}
    return JSONResponse({'status': 'ok', 'message': 'FP cache cleared'})

# ── Unified Reload Master — โหลดใหม่ทั้ง 3 ไฟล์ ─────────────────────────────
@app.get("/api/reload-master")
async def api_reload_master_all():
    """โหลด Master data ใหม่ทั้งหมด:
       1. Master data HEI.xlsx  (items/spec for Expander, HP, Cutting)
       2. Data Man MC Mat.xlsx  (machines, checkers, employees)
       3. Master speclist.xlsm  (FP spec list)
    """
    global _fp_spec_raw, _fp_item_rm_cache, _fp_sheet_cache
    results = {}
    errors  = []

    # ── 1. Master data HEI.xlsx ──────────────────────────────────────
    try:
        n = build_items_json()
        results['master_hei'] = f'✓ Master data HEI.xlsx — {n} items'
    except Exception as e:
        err = f'✗ Master data HEI.xlsx: {e}'
        results['master_hei'] = err; errors.append(err)

    # ── 2. Data Man MC Mat.xlsx ──────────────────────────────────────
    try:
        build_cutting_people()
        results['data_man'] = '✓ Data Man MC Mat.xlsx — machines / checkers / employees'
    except Exception as e:
        err = f'✗ Data Man MC Mat.xlsx: {e}'
        results['data_man'] = err; errors.append(err)

    # ── 3. Master speclist.xlsm (FP) ────────────────────────────────
    try:
        _fp_spec_raw = None; _fp_item_rm_cache = None; _fp_sheet_cache = {}
        rows, load_err = _fp_load_speclist()   # โหลดทันที (ไม่รอ lazy)
        if load_err:
            raise Exception(load_err)
        results['speclist'] = f'✓ Master speclist.xlsm — {len(rows)} rows'
    except Exception as e:
        err = f'✗ Master speclist.xlsm: {e}'
        results['speclist'] = err; errors.append(err)

    # ── 4. Hairpin Insert caches (HEI / speclist / people) ──────────
    try:
        _hpins_store['data']     = _hpins_load_data()
        _hpins_store['hei']      = _hpins_load_hei()
        _hpins_store['speclist'] = _hpins_load_speclist()
        results['hp_insert'] = (f"✓ Hairpin Insert — HEI {len(_hpins_store['hei'])} / "
                                f"Speclist {len(_hpins_store['speclist'])} items")
    except Exception as e:
        err = f'✗ Hairpin Insert: {e}'
        results['hp_insert'] = err; errors.append(err)

    return JSONResponse({
        'ok':      len(errors) == 0,
        'results': results,
        'errors':  errors,
        'message': ('โหลด Master ทั้ง 3 ไฟล์สำเร็จ' if not errors
                    else f'โหลดสำเร็จบางส่วน ({len(errors)} error)'),
    })

@app.get("/fp/api/machine")
async def fp_api_machine(q: str = ''):
    if not q: return JSONResponse({'found': False, 'data': None, 'name': ''})
    row, name, err = _fp_find_match('machine', q)
    if row: return JSONResponse({'found': True, 'data': row, 'name': name})
    return JSONResponse({'found': False, 'error': err, 'query': q, 'name': ''})

@app.get("/fp/api/checker")
async def fp_api_checker(q: str = ''):
    if not q: return JSONResponse({'found': False, 'data': None, 'name': ''})
    row, name, err = _fp_find_match('checker', q)
    if row: return JSONResponse({'found': True, 'data': row, 'name': name})
    return JSONResponse({'found': False, 'error': err, 'query': q, 'name': ''})

@app.get("/fp/api/confirm")
async def fp_api_confirm(q: str = ''):
    if not q: return JSONResponse({'found': False, 'data': None, 'name': ''})
    row, name, err = _fp_find_match('confirm', q)
    if row: return JSONResponse({'found': True, 'data': row, 'name': name})
    return JSONResponse({'found': False, 'error': err, 'query': q, 'name': ''})

@app.get("/fp/api/speclist")
async def fp_api_speclist(q: str = ''):
    if not q: return JSONResponse({'found': False, 'fields': None})
    if _fp_spec_raw is None: return JSONResponse({'found': False, 'loading': True, 'error': 'loading'})
    fields, err = _fp_find_spec(q)
    if fields: return JSONResponse({'found': True, 'fields': fields})
    return JSONResponse({'found': False, 'error': err, 'query': q})

@app.get("/fp/api/hei_item")
async def fp_api_hei_item(q: str = ''):
    if not q: return JSONResponse({'found': False, 'fields': None})
    fields, err = _fp_find_hei(q)
    if fields: return JSONResponse({'found': True, 'fields': fields})
    return JSONResponse({'found': False, 'error': err, 'query': q})

@app.get("/fp/api/item_rm_lookup")
async def fp_api_item_rm(q: str = ''):
    if not q: return JSONResponse({'found': False, 'spec': ''})
    cache, err = _fp_load_item_rm()
    if err or not cache: return JSONResponse({'found': False, 'spec': '', 'error': err or 'not loaded'})
    spec = cache.get(q.strip().upper(), '')
    return JSONResponse({'found': bool(spec), 'spec': spec})

# ── FP Record API ──────────────────────────────────────────────────────────────
@app.post("/fp/api/record")
async def fp_api_save(request: Request):
    d = await request.json()
    try:
        now = datetime.now()
        dv  = d.get('date','') or now.strftime('%Y-%m-%d')
        tv  = d.get('time','') or now.strftime('%H:%M')
        con = _fp_get_db()
        cur = con.execute('''
            INSERT INTO fp_records (
                date, time, mc, prod_month, lot, line, seq,
                type, item, desc1, model,
                vendor, coil_no, pack_no, al_item, mat_type,
                appear_coil, appear_flare, appear_slit, result_appear,
                fp1, fp2, fp3, fp4, fp_avg, result_fp,
                ff1, ff2, ff3, ff4, ff_avg, result_ff,
                qty_fin, result_qty, fin_die, holes,
                final, checker, confirmer,
                spec_pitch, oil, spec_thick, compare_al, sc_appear,
                sc1, sc2, sc3, sc4, sc_avg, result_sc,
                order_prod, saved_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            dv, tv,
            d.get('mc_no',''), d.get('prod_month',''), d.get('lot',''),
            d.get('line',''),  d.get('seq',''),
            d.get('type',''),  d.get('item',''),
            d.get('desc1',''), d.get('model',''),
            d.get('vendor',''), d.get('coil_no',''), d.get('pack_no',''),
            d.get('al_item',''), d.get('mat_type',''),
            d.get('appear_coil',''), d.get('appear_flare',''),
            d.get('appear_slit',''), d.get('result_appear',''),
            d.get('fp1',''), d.get('fp2',''), d.get('fp3',''), d.get('fp4',''),
            d.get('fp_avg',''), d.get('result_fp',''),
            d.get('ff1',''), d.get('ff2',''), d.get('ff3',''), d.get('ff4',''),
            d.get('ff_avg',''), d.get('result_ff',''),
            d.get('qty_fin',''), d.get('result_qty',''),
            d.get('fin_die',''), d.get('holes',''),
            d.get('final',''), d.get('checker',''), d.get('confirmer',''),
            d.get('spec_pitch',''), d.get('oil',''), d.get('spec_thick',''),
            d.get('compare_al',''), d.get('sc_appear',''),
            d.get('sc1',''), d.get('sc2',''), d.get('sc3',''), d.get('sc4',''),
            d.get('sc_avg',''), d.get('result_sc',''),
            d.get('order_prod',''), now.isoformat()
        ))
        con.commit()
        rid = cur.lastrowid; con.close()
        return JSONResponse({'ok': True, 'row': rid})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.post("/fp/api/export-excel")
async def fp_api_export_excel(request: Request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from fastapi.responses import StreamingResponse
        import io

        body = await request.json()
        records = body.get('ids', [])   # list of record dicts from frontend

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'FP First Lot'

        headers = [
            'No','Date','Time','Machine','Item FG','Type',
            'Lot','Line','Seq',
            'Vendor','Pack No','Mat Type','Oil Type',
            'Appear Coil','Appear Flare','Appear Slit','Appear Result',
            'FP1','FP2','FP3','FP4','FP Avg','FP Result','Spec Pitch',
            'FF1','FF2','FF3','FF4','FF Avg','FF Result',
            "Q'ty Fin","Q'ty Result",
            'SC1','SC2','SC3','SC4','SC Avg','SC Result',
            'Checker','Confirmer','Final'
        ]
        ws.append(headers)

        thin = Side(style='thin', color='B0BEC5')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='B45309')
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 20

        pf = PatternFill('solid', fgColor='C8E6C9')
        nf = PatternFill('solid', fgColor='FFCDD2')

        for i, r in enumerate(records, 1):
            ws.append([
                i,
                r.get('date',''),   r.get('time',''),   r.get('mc',''),
                r.get('item',''),   r.get('type',''),
                r.get('lot',''),    r.get('line',''),    r.get('seq',''),
                r.get('vendor',''), r.get('pack_no',''), r.get('mat_type',''), r.get('oil',''),
                r.get('appear_coil',''), r.get('appear_flare',''),
                r.get('appear_slit',''), r.get('result_appear',''),
                r.get('fp1',''),  r.get('fp2',''),  r.get('fp3',''),  r.get('fp4',''),
                r.get('fp_avg',''), r.get('result_fp',''), r.get('spec_pitch',''),
                r.get('ff1',''),  r.get('ff2',''),  r.get('ff3',''),  r.get('ff4',''),
                r.get('ff_avg',''), r.get('result_ff',''),
                r.get('qty_fin',''), r.get('result_qty',''),
                r.get('sc1',''),  r.get('sc2',''),  r.get('sc3',''),  r.get('sc4',''),
                r.get('sc_avg',''), r.get('result_sc',''),
                r.get('checker',''), r.get('confirmer',''), r.get('final','')
            ])
            ri = ws.max_row
            for cell in ws[ri]:
                cell.border = bdr
                cell.alignment = Alignment(vertical='center')
            fin = str(r.get('final',''))
            ws.cell(ri, len(headers)).fill = (
                pf if fin == 'OK' else (nf if fin == 'NG' else PatternFill()))

        col_widths = [
            5, 11, 7, 22, 16, 8,
            10, 6, 6,
            16, 14, 14, 12,
            10, 10, 10, 10,
            7, 7, 7, 7, 8, 9, 10,
            7, 7, 7, 7, 8, 9,
            8, 9,
            7, 7, 7, 7, 8, 9,
            18, 18, 7
        ]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'fp_firstlot_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return StreamingResponse(
            buf,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/fp/api/records")
async def fp_api_records(request: Request):
    from datetime import date as _d
    p   = dict(request.query_params)
    q_t = p.get('q','').strip().upper()
    mc_f= p.get('mc','').strip()
    st  = p.get('status','').strip().upper()
    fr  = p.get('from','').strip()
    to_ = p.get('to','').strip()
    lim = max(1, int(p.get('limit','50')))
    pg  = max(1, int(p.get('page','1')))
    today = _d.today().isoformat()
    eff = "COALESCE(NULLIF(date,''), substr(saved_at,1,10))"

    con = _fp_get_db()
    conds, vals = [], []
    if q_t:
        like = f'%{q_t}%'
        conds.append('(mc LIKE ? OR item LIKE ? OR lot LIKE ? OR checker LIKE ?)')
        vals += [like, like, like, like]
    if mc_f: conds.append('mc=?'); vals.append(mc_f)
    if st == 'OK':  conds.append("final='OK'")
    elif st == 'NG': conds.append("final='NG'")
    if fr:  conds.append(f'{eff}>=?'); vals.append(fr)
    if to_: conds.append(f'{eff}<=?'); vals.append(to_)
    where = ('WHERE ' + ' AND '.join(conds)) if conds else ''

    total_all = con.execute('SELECT COUNT(*) FROM fp_records').fetchone()[0]
    ok_all    = con.execute("SELECT COUNT(*) FROM fp_records WHERE final='OK'").fetchone()[0]
    ng_all    = con.execute("SELECT COUNT(*) FROM fp_records WHERE final='NG'").fetchone()[0]
    today_cnt = con.execute(f'SELECT COUNT(*) FROM fp_records WHERE {eff}=?', (today,)).fetchone()[0]
    mc_active = con.execute(f'SELECT COUNT(DISTINCT mc) FROM fp_records WHERE {eff}=?', (today,)).fetchone()[0]
    machines  = [r[0] for r in con.execute("SELECT DISTINCT mc FROM fp_records WHERE mc!='' ORDER BY mc").fetchall()]

    mc_status = []
    for m in machines:
        r = con.execute('SELECT date,time,line,seq,item,type,final,checker,saved_at FROM fp_records WHERE mc=? ORDER BY id DESC LIMIT 1', (m,)).fetchone()
        if r:
            rd2 = dict(r); s2 = rd2.get('saved_at','')
            rd2['date'] = rd2.get('date','') or s2[:10]
            rd2['time'] = rd2.get('time','') or s2[11:16]
            mc_status.append({'mc': m, **rd2})

    total = con.execute(f'SELECT COUNT(*) FROM fp_records {where}', vals).fetchone()[0]
    pages = max(1, (total + lim - 1) // lim)
    pg    = min(pg, pages)
    rows  = con.execute(
        f'SELECT * FROM fp_records {where} ORDER BY id DESC LIMIT ? OFFSET ?',
        vals + [lim, (pg-1)*lim]
    ).fetchall()
    con.close()

    records = []
    for i, r in enumerate(rows, (pg-1)*lim + 1):
        rd = dict(r); sv = rd.get('saved_at','')
        records.append({
            'no': i,
            'date': rd.get('date','') or sv[:10],
            'time': rd.get('time','') or sv[11:16],
            'mc': rd.get('mc',''), 'prod_month': rd.get('prod_month',''),
            'lot': rd.get('lot',''), 'line': rd.get('line',''), 'seq': rd.get('seq',''),
            'type': rd.get('type',''), 'item': rd.get('item',''),
            'vendor': rd.get('vendor',''), 'coil_no': rd.get('coil_no',''),
            'pack_no': rd.get('pack_no',''),
            'al_item': rd.get('al_item',''), 'mat_type': rd.get('mat_type',''),
            'appear_coil': rd.get('appear_coil',''), 'result_appear': rd.get('result_appear',''),
            'sc_appear': rd.get('sc_appear',''),
            'fp1': rd.get('fp1',''), 'fp2': rd.get('fp2',''),
            'fp3': rd.get('fp3',''), 'fp4': rd.get('fp4',''),
            'fp_avg': rd.get('fp_avg',''), 'result_fp': rd.get('result_fp',''),
            'spec_pitch': rd.get('spec_pitch',''),
            'ff1': rd.get('ff1',''), 'ff2': rd.get('ff2',''),
            'ff3': rd.get('ff3',''), 'ff4': rd.get('ff4',''),
            'ff_avg': rd.get('ff_avg',''), 'result_ff': rd.get('result_ff',''),
            'qty_fin': rd.get('qty_fin',''), 'result_qty': rd.get('result_qty',''),
            'sc1': rd.get('sc1',''), 'sc2': rd.get('sc2',''),
            'sc3': rd.get('sc3',''), 'sc4': rd.get('sc4',''),
            'sc_avg': rd.get('sc_avg',''), 'result_sc': rd.get('result_sc',''),
            'checker': rd.get('checker',''), 'confirmer': rd.get('confirmer',''),
            'oil': rd.get('oil',''),
            'final': rd.get('final',''),
            'saved_at': rd.get('saved_at',''),
        })

    return JSONResponse({
        'total': total, 'total_all': total_all,
        'ok_all': ok_all, 'ng_all': ng_all,
        'today': today_cnt, 'mc_active': mc_active,
        'page': pg, 'pages': pages, 'limit': lim,
        'machines': machines, 'mc_status': mc_status,
        'records': records,
    })

# ── FP Admin API ───────────────────────────────────────────────────────────────
@app.get("/fp/api/records/all")
async def fp_api_records_all():
    con = _fp_get_db()
    rows = con.execute('SELECT * FROM fp_records ORDER BY id DESC').fetchall()
    con.close()
    result = []
    for r in rows:
        rd = dict(r); sv = rd.get('saved_at','')
        result.append({
            'id':        rd['id'],
            'date':      rd.get('date','') or sv[:10],
            'time':      rd.get('time','') or sv[11:16],
            'mc':        rd.get('mc',''),
            'type':      rd.get('type',''),
            'item':      rd.get('item',''),
            'lot':       rd.get('lot',''),
            'line':      rd.get('line',''),
            'seq':       rd.get('seq',''),
            'vendor':    rd.get('vendor',''),
            'oil':       rd.get('oil',''),
            'fp_avg':    rd.get('fp_avg',''),
            'ff_avg':    rd.get('ff_avg',''),
            'result_fp': rd.get('result_fp',''),
            'result_ff': rd.get('result_ff',''),
            'qty_fin':   rd.get('qty_fin',''),
            'result_qty':rd.get('result_qty',''),
            'checker':   rd.get('checker',''),
            'confirmer': rd.get('confirmer',''),
            'final':     rd.get('final',''),
        })
    return JSONResponse(result)

@app.put("/fp/api/record/{rid}")
async def fp_api_update(rid: int, request: Request):
    try:
        d = await request.json()
        allowed = ['date','time','mc','lot','line','seq','type',
                   'vendor','oil','checker','confirmer','final']
        sets = ', '.join(f'{c}=?' for c in allowed)
        vals = [d.get(c,'') for c in allowed] + [rid]
        con = _fp_get_db()
        con.execute(f'UPDATE fp_records SET {sets} WHERE id=?', vals)
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.delete("/fp/api/record/{rid}")
async def fp_api_delete(rid: int):
    try:
        con = _fp_get_db()
        con.execute('DELETE FROM fp_records WHERE id=?', (rid,))
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.post("/fp/api/records/bulk-delete")
async def fp_api_bulk_delete(request: Request):
    try:
        ids = (await request.json()).get('ids', [])
        if not ids: return JSONResponse({'ok': False, 'error': 'No IDs provided'})
        con = _fp_get_db()
        placeholders = ','.join('?' * len(ids))
        con.execute(f'DELETE FROM fp_records WHERE id IN ({placeholders})', ids)
        con.commit(); con.close()
        return JSONResponse({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

# ── Shift helpers (Day 07:40-19:39 / Night 19:40-07:39) ───────────────────────
def _hei_shift_start(now=None):
    now = now or datetime.now()
    t = now.hour * 60 + now.minute
    DAY, NIGHT = 7 * 60 + 40, 19 * 60 + 40
    if DAY <= t < NIGHT:
        return now.replace(hour=7, minute=40, second=0, microsecond=0)
    elif t >= NIGHT:
        return now.replace(hour=19, minute=40, second=0, microsecond=0)
    else:
        return (now - timedelta(days=1)).replace(hour=19, minute=40, second=0, microsecond=0)

def _hei_shift_name(now=None):
    now = now or datetime.now()
    t = now.hour * 60 + now.minute
    return 'day' if (7*60+40) <= t < (19*60+40) else 'night'

def _hei_parse_dt(date_val, time_val=None):
    """Parse YYYY-MM-DD / DD/MM/YYYY / ISO 'T' datetime (+ optional HH:MM).
    Auto-converts Thai Buddhist Era years (≥ 2500) → Gregorian by subtracting 543.
    """
    if not date_val:
        return None
    s = str(date_val)
    try:
        if 'T' in s:
            dt = datetime.fromisoformat(s[:19])
            # Fromisoformat may give BE year if stored that way
            if dt.year >= 2500:
                dt = dt.replace(year=dt.year - 543)
            return dt
        if '/' in s:                       # DD/MM/YYYY
            d, mo, y = [int(x) for x in s.split('/')[:3]]
        elif '-' in s:                     # YYYY-MM-DD
            y, mo, d = [int(x) for x in s.split('-')[:3]]
        else:
            return None
        if y >= 2500:                      # Thai Buddhist Era → Gregorian
            y -= 543
        hh = mm = 0
        if time_val and ':' in str(time_val):
            p = str(time_val).split(':')
            hh, mm = int(p[0]), int(p[1])
        return datetime(y, mo, d, hh, mm)
    except Exception:
        return None

# ── Machine Records API — table of CURRENT-SHIFT records for one machine ───────
@app.get("/api/machine-records/{mtype}/{mnum}")
async def api_machine_records(mtype: str, mnum: int):
    """All records of one machine within the CURRENT shift.

    Columns returned: checker (ID), line, seq, item
      item = Item FG by default; for HP process item = Item Hairpin.

    Usage: /api/machine-records/fp/20   → Finpress No.20
           /api/machine-records/ex/17   → Expander No.17
           /api/machine-records/hp/19   → Hairpin Bender No.19
           /api/machine-records/ct/13   → Cutting No.13
    """
    mtype = mtype.lower()
    pad   = str(mnum).zfill(2)
    start = _hei_shift_start()
    out, label = [], ''

    if mtype == 'fp':
        con = _fp_get_db()
        rows = con.execute(
            'SELECT date,time,line,seq,item,checker,saved_at FROM fp_records '
            'WHERE mc LIKE ? ORDER BY id DESC LIMIT 300', (f'%no.{pad}%',)
        ).fetchall(); con.close()
        label = f'Finpress No.{pad}'
        for r in rows:
            rd = dict(r)
            dt = _hei_parse_dt(rd.get('date') or rd.get('saved_at'), rd.get('time'))
            if dt and dt >= start:
                out.append({'date': dt.strftime('%Y-%m-%d'), 'time': dt.strftime('%H:%M'),
                            'checker': rd.get('checker','') or '',
                            'line': rd.get('line','') or '', 'seq': rd.get('seq','') or '',
                            'item': rd.get('item','') or ''})

    elif mtype == 'ex':
        con = get_db()
        rows = con.execute(
            'SELECT date,time,line,seq,item_fg,checker FROM records '
            'WHERE machine LIKE ? ORDER BY id DESC LIMIT 300', (f'EX{pad}%',)
        ).fetchall(); con.close()
        label = f'Expander No.{mnum}'
        for r in rows:
            rd = dict(r)
            dt = _hei_parse_dt(rd.get('date'), rd.get('time'))
            if dt and dt >= start:
                out.append({'date': dt.strftime('%Y-%m-%d'), 'time': dt.strftime('%H:%M'),
                            'checker': rd.get('checker','') or '',
                            'line': rd.get('line','') or '', 'seq': rd.get('seq','') or '',
                            'item': rd.get('item_fg','') or ''})

    elif mtype == 'hp':
        con = hp_get_db()
        rows = con.execute(
            'SELECT ts,line,seq,item,item_fg,checker,saved_at FROM hp_records '
            'WHERE mc LIKE ? ORDER BY id DESC LIMIT 300', (f'HB{pad}',)
        ).fetchall(); con.close()
        label = f'Hairpin Bender No.{pad}'
        for r in rows:
            rd = dict(r)
            # saved_at = LOCAL time (matches dashboard); ts = UTC 'Z' → don't use for shift compare
            dt = _hei_parse_dt(rd.get('saved_at') or rd.get('ts'))
            if dt and dt >= start:
                # HP process → Item shows Hairpin item, not FG
                out.append({'date': dt.strftime('%Y-%m-%d'), 'time': dt.strftime('%H:%M'),
                            'checker': rd.get('checker','') or '',
                            'line': rd.get('line','') or '', 'seq': rd.get('seq','') or '',
                            'item': rd.get('item','') or ''})

    elif mtype == 'ct':
        con = cutting_get_db()
        rows = con.execute(
            'SELECT saved_at,line,seq,tag_fg_no1,id_card FROM cutting_records '
            'WHERE mc_line LIKE ? ORDER BY id DESC LIMIT 300', (f'Cutting no.{pad}',)
        ).fetchall(); con.close()
        label = f'Cutting No.{pad}'
        for r in rows:
            rd = dict(r)
            dt = _hei_parse_dt(rd.get('saved_at'))
            if dt and dt >= start:
                out.append({'date': dt.strftime('%Y-%m-%d'), 'time': dt.strftime('%H:%M'),
                            'checker': rd.get('id_card','') or '',
                            'line': rd.get('line','') or '', 'seq': rd.get('seq','') or '',
                            'item': rd.get('tag_fg_no1','') or ''})
    elif mtype == 'ov':
        con = oven_get_db()
        rows = con.execute(
            'SELECT date,time,line,seq,item_fg,checker_id,machine_code,saved_at FROM oven_records '
            'WHERE machine_code LIKE ? ORDER BY id DESC LIMIT 300', (f'%{pad}%',)
        ).fetchall(); con.close()
        label = f'Oven Oil No.{pad}'
        for r in rows:
            rd = dict(r)
            dt = _hei_parse_dt(rd.get('date') or rd.get('saved_at'), rd.get('time'))
            if dt and dt >= start:
                out.append({'date': dt.strftime('%Y-%m-%d'), 'time': dt.strftime('%H:%M'),
                            'checker': rd.get('checker_id','') or '',
                            'line': rd.get('line','') or '', 'seq': rd.get('seq','') or '',
                            'item': rd.get('item_fg','') or ''})
    else:
        return JSONResponse({'ok': False, 'error': 'unknown type'}, status_code=400)

    return JSONResponse({
        'ok': True, 'machine': label, 'shift': _hei_shift_name(),
        'count': len(out), 'records': out,
    })


# ══════════════════════════════════════════════════════════════════════════════
# HAIRPIN INSERT FIRST LOT (CH-541-04-01)  —  /hp_insert/firstlot
# ══════════════════════════════════════════════════════════════════════════════
_HPINS_DB         = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Server_firstlot\hp_insert_records.db"
_HPINS_SPECLIST   = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Master speclist.xlsm"
_HPINS_HEI_SHEET  = 'Master Review-2025 (2)'
_HPINS_SPEC_SHEET = 'Speclistnew'

_hpins_store = {'data': None, 'hei': None, 'speclist': None}

def _hpins_init_db():
    con = sqlite3.connect(_HPINS_DB)
    con.execute('''CREATE TABLE IF NOT EXISTS records (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        rec_date         TEXT,
        rec_time         TEXT,
        machine_finpress TEXT,
        machine_bender   TEXT,
        machine_bender2  TEXT,
        item_wip         TEXT,
        prod_type        TEXT,
        item_hp1         TEXT,
        item_hp2         TEXT,
        hp_tube_type     TEXT,
        hp_tube_type2    TEXT,
        line             TEXT,
        seq              TEXT,
        prod_month       TEXT,
        lot              TEXT,
        item_fin         TEXT,
        fin_pitch        TEXT,
        fin_type         TEXT,
        mark_color       TEXT,
        spec_hairpin     TEXT,
        spec_fin         TEXT,
        spec_assembly    TEXT,
        spec_fin_pattern TEXT,
        spec_appearance  TEXT,
        name_checker     TEXT,
        name_confirm     TEXT,
        final_check      TEXT,
        created_at       TEXT DEFAULT (datetime('now','localtime'))
    )''')
    cols = [r[1] for r in con.execute('PRAGMA table_info(records)').fetchall()]
    for c in ('prod_month', 'item_fin', 'fin_pitch', 'fin_type', 'machine_bender2', 'mark_color', 'verified',
              'hp1_line', 'hp1_seq', 'hp1_prod_month', 'hp2_line', 'hp2_seq', 'hp2_prod_month'):
        if c not in cols:
            con.execute(f'ALTER TABLE records ADD COLUMN {c} TEXT')
    con.commit(); con.close()

_hpins_init_db()

def _hpins_load_data():
    """Data Man MC Mat.xlsx → checkers / confirms / machines"""
    import openpyxl
    wb = openpyxl.load_workbook(PEOPLE_XLSX, read_only=True, data_only=True)
    checkers = []
    for row in wb['Name Checker'].iter_rows(min_row=3, values_only=True):
        full_id  = str(row[0] or '').strip()
        name     = str(row[2] or '').strip()
        short_id = full_id[5:] if len(full_id) > 5 else full_id
        if short_id and full_id:
            checkers.append({'id': short_id, 'full_id': full_id, 'name': name})
    confirms = []
    for row in wb['Name confirm'].iter_rows(min_row=3, values_only=True):
        raw_id = row[0]
        if raw_id is None:
            continue
        short_id = str(int(raw_id)) if isinstance(raw_id, (int, float)) else str(raw_id).strip()
        name = str(row[2] or '').strip()
        role = str(row[3] or '').strip() if len(row) > 3 else ''
        if short_id:
            confirms.append({'id': short_id, 'name': name, 'role': role})
    machines = []
    for row in wb['Machine'].iter_rows(min_row=2, values_only=True):
        code  = str(row[0] or '').strip()
        mname = str(row[1] or '').strip() if len(row) > 1 else ''
        if code:
            machines.append({'code': code, 'name': mname})
    wb.close()
    return {'checkers': checkers, 'confirms': confirms, 'machines': machines}

def _hpins_load_hei():
    """Master data HEI.xlsx: key=Col F, hp1=AE, hp2=AK, finindex=AX, finpitch=AZ"""
    import openpyxl
    wb  = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    ws  = wb[_HPINS_HEI_SHEET]
    hei = {}
    # Mark Color columns BC:BJ (idx 54-61) — each filled cell = a valid color
    MARK_COLOR_COLS = [
        (54, 'GREEN'), (55, 'BLUE'),  (56, 'RED'),    (57, 'WHITE'),
        (58, 'PURPLE'),(59, 'BLACK'), (60, 'YELLOW'), (61, 'ORANGE'),
    ]
    # Copper raw-material columns → pipe diameter
    PIPE7_COLS = (25, 26, 27)   # Z/AA/AB = COPPER 7x...
    PIPE5_COLS = (28, 29)       # AC/AD   = COPPER 5x...
    def _cell(row, i): return str(row[i] or '').strip() if i < len(row) else ''
    for row in ws.iter_rows(min_row=4, values_only=True):
        item     = str(row[5]  or '').strip()
        hp1      = str(row[30] or '').strip()
        hp2      = str(row[36] or '').strip()
        finindex = str(row[49] or '').strip()
        finpitch = str(row[51] or '').strip()
        markcolors = [name for idx, name in MARK_COLOR_COLS
                      if idx < len(row) and str(row[idx] or '').strip()]
        has5 = any(_cell(row, i) for i in PIPE5_COLS)
        has7 = any(_cell(row, i) for i in PIPE7_COLS)
        pipe = '5' if (has5 and not has7) else ('7' if (has7 and not has5) else '')
        if item:
            hei[item.upper()] = {'hp1': hp1, 'hp2': hp2,
                                 'finpitch': finpitch, 'finindex': finindex,
                                 'markcolors': markcolors, 'pipe': pipe}
    wb.close()
    return hei

def _hpins_normalize_fintype(raw):
    r = (raw or '').strip().upper()
    if r.startswith('PRECOAT'):
        return 'PRECOAT'
    if r.startswith('HN'):
        return 'HN'
    return (raw or '').strip()

def _hpins_load_speclist():
    """Master speclist.xlsm: key=Col C, fintype=Col AN (normalized)"""
    import openpyxl
    wb   = openpyxl.load_workbook(_HPINS_SPECLIST, read_only=True, data_only=True)
    ws   = wb[_HPINS_SPEC_SHEET]
    spec = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        item   = str(row[2]  or '').strip()
        raw_ft = str(row[39] or '').strip()
        if item:
            spec[item.upper()] = _hpins_normalize_fintype(raw_ft)
    wb.close()
    return spec

def _hpins_ensure(key, loader):
    """Lazy-load master data on first request (keeps server startup fast)."""
    if _hpins_store[key] is None:
        try:
            _hpins_store[key] = loader()
        except Exception as e:
            print(f'[HP Insert] load {key} ERROR: {e}')
            _hpins_store[key] = {} if key != 'data' else {'checkers': [], 'confirms': [], 'machines': []}
    return _hpins_store[key]

# ── Pages ──────────────────────────────────────────────────────────────────────
@app.get("/hp_insert/firstlot")
@app.get("/hp_insert/firstlot/")
async def page_hpins():
    return FileResponse(os.path.join(_FL_DIR, "hp_insert.html"))

# ── API ────────────────────────────────────────────────────────────────────────
@app.get("/hp_insert/api/data")
async def hpins_api_data():
    return JSONResponse(_hpins_ensure('data', _hpins_load_data))

@app.get("/hp_insert/api/hei_lookup")
async def hpins_api_hei(item: str = ''):
    item = (item or '').strip().upper()
    if not item:
        return JSONResponse({'found': False, 'hp1': '', 'hp2': '', 'finpitch': ''})
    row = _hpins_ensure('hei', _hpins_load_hei).get(item)
    if row:
        return JSONResponse({'found': True, 'hp1': row['hp1'], 'hp2': row['hp2'],
                             'finpitch': row.get('finpitch', ''),
                             'finindex': row.get('finindex', ''),
                             'markcolors': row.get('markcolors', []),
                             'pipe': row.get('pipe', '')})
    return JSONResponse({'found': False, 'hp1': '', 'hp2': '', 'finpitch': '', 'markcolors': [], 'pipe': ''})

@app.get("/hp_insert/api/speclist_lookup")
async def hpins_api_spec(item: str = ''):
    item = (item or '').strip().upper()
    if not item:
        return JSONResponse({'found': False, 'fintype': ''})
    ft = _hpins_ensure('speclist', _hpins_load_speclist).get(item)
    if ft is not None:
        return JSONResponse({'found': True, 'fintype': ft})
    return JSONResponse({'found': False, 'fintype': ''})

@app.get("/hp_insert/api/jpeg")
async def hpins_api_jpeg(item: str = ''):
    item = (item or '').strip()
    if not item:
        return JSONResponse({'error': 'missing item'}, status_code=400)
    for fn in (f'{item}-INS.jpg', f'{item}-INS.JPG'):
        path = os.path.join(_PC_DIR, fn)
        if os.path.isfile(path):
            return FileResponse(path, media_type='image/jpeg')
    return JSONResponse({'error': 'not found'}, status_code=404)

@app.get("/hp_insert/api/reload")
async def hpins_api_reload():
    try:
        _hpins_store['data']     = _hpins_load_data()
        _hpins_store['hei']      = _hpins_load_hei()
        _hpins_store['speclist'] = _hpins_load_speclist()
        return JSONResponse({'ok': True,
                             'checkers': len(_hpins_store['data']['checkers']),
                             'machines': len(_hpins_store['data']['machines']),
                             'hei':      len(_hpins_store['hei']),
                             'speclist': len(_hpins_store['speclist'])})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)})

@app.get("/hp_insert/api/records")
async def hpins_api_records():
    con = sqlite3.connect(_HPINS_DB)
    con.row_factory = sqlite3.Row
    rows = [dict(r) for r in con.execute(
        'SELECT * FROM records ORDER BY id DESC').fetchall()]
    con.close()
    return JSONResponse(rows)

@app.get("/hp_insert/api/records/delete")
async def hpins_api_delete(id: str = ''):
    if not id:
        return JSONResponse({'ok': False, 'error': 'missing id'})
    con = sqlite3.connect(_HPINS_DB)
    con.execute('DELETE FROM records WHERE id=?', (id,))
    con.commit(); con.close()
    return JSONResponse({'ok': True})

@app.get("/hp_insert/api/records/verify")
async def hpins_api_verify(id: str = '', value: str = '1'):
    if not id:
        return JSONResponse({'ok': False, 'error': 'missing id'})
    con = sqlite3.connect(_HPINS_DB)
    con.execute('UPDATE records SET verified=? WHERE id=?', ('1' if value == '1' else '', id))
    con.commit(); con.close()
    return JSONResponse({'ok': True})

@app.post("/hp_insert/api/save")
async def hpins_api_save(request: Request):
    try:
        d = await request.json()
        con = sqlite3.connect(_HPINS_DB)
        con.execute('''INSERT INTO records
            (rec_date,rec_time,machine_finpress,machine_bender,machine_bender2,item_wip,prod_type,
             item_hp1,item_hp2,hp1_line,hp1_seq,hp1_prod_month,hp2_line,hp2_seq,hp2_prod_month,
             hp_tube_type,hp_tube_type2,line,seq,prod_month,lot,
             item_fin,fin_pitch,fin_type,mark_color,
             spec_hairpin,spec_fin,spec_assembly,spec_fin_pattern,spec_appearance,
             name_checker,name_confirm,final_check)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            d.get('rec_date',''), d.get('rec_time',''),
            d.get('machine_finpress',''), d.get('machine_bender',''),
            d.get('machine_bender2',''),
            d.get('item_wip',''), d.get('prod_type',''),
            d.get('item_hp1',''), d.get('item_hp2',''),
            d.get('hp1_line',''), d.get('hp1_seq',''), d.get('hp1_prod_month',''),
            d.get('hp2_line',''), d.get('hp2_seq',''), d.get('hp2_prod_month',''),
            d.get('hp_tube_type',''), d.get('hp_tube_type2',''),
            d.get('line',''), d.get('seq',''),
            d.get('prod_month',''), d.get('lot',''),
            d.get('item_fin',''), d.get('fin_pitch',''),
            d.get('fin_type',''), d.get('mark_color',''),
            d.get('spec_hairpin',''), d.get('spec_fin',''),
            d.get('spec_assembly',''), d.get('spec_fin_pattern',''),
            d.get('spec_appearance',''),
            d.get('name_checker',''), d.get('name_confirm',''),
            d.get('final_check','')
        ))
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


# ══════════════════════════════════════════════════════════════════════════════
# OVEN OIL FIRST LOT  —  /oven/firstlot
#   Temp-sensor check (Set vs Actual ± tolerance) recorded per time-slot.
# ══════════════════════════════════════════════════════════════════════════════
_OVEN_DB = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\Server_firstlot\oven_records.db"

def oven_get_db():
    con = sqlite3.connect(_OVEN_DB)
    con.row_factory = sqlite3.Row
    return con

def oven_init_db():
    con = sqlite3.connect(_OVEN_DB)
    con.execute('''CREATE TABLE IF NOT EXISTS oven_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT, time TEXT,
        shift TEXT, slot TEXT,
        checker TEXT, checker_id TEXT, confirm_checker TEXT,
        machine TEXT, machine_code TEXT,
        line TEXT, seq TEXT, lot TEXT,
        item_fg TEXT, item_type TEXT,
        set_temp TEXT, tol TEXT, actual_temp TEXT, temp_diff TEXT, temp_result TEXT,
        remark TEXT, status TEXT,
        saved_at TEXT
    )''')
    con.commit()
    # Migrate older tables — add any missing columns
    existing = {r[1] for r in con.execute("PRAGMA table_info(oven_records)").fetchall()}
    for col in ['shift','slot','machine_code','temp_diff','temp_result','lot']:
        if col not in existing:
            con.execute(f'ALTER TABLE oven_records ADD COLUMN {col} TEXT')
    con.commit(); con.close()

oven_init_db()

# ── Pages ──────────────────────────────────────────────────────────────────────
@app.get("/oven/firstlot")
@app.get("/oven/firstlot/")
async def page_oven():
    return FileResponse(os.path.join(_FL_DIR, "oven.html"))

# ── Static (shared public/ data + images) ─────────────────────────────────────
@app.get("/oven/data/{filename:path}")
async def oven_static_data(filename: str):
    path = os.path.join(_FL_DIR, "data", filename)
    if os.path.exists(path):
        return FileResponse(path, headers={"Cache-Control": "no-store, no-cache, must-revalidate"})
    return JSONResponse({"error": "not found"}, status_code=404)

@app.get("/oven/images/{filename:path}")
async def oven_static_images(filename: str):
    path = os.path.join(_FL_DIR, "images", filename)
    return FileResponse(path) if os.path.exists(path) else JSONResponse({"error": "not found"}, status_code=404)

# ── API ────────────────────────────────────────────────────────────────────────
@app.get("/oven/api/records")
async def oven_api_records():
    con = oven_get_db()
    rows = [dict(r) for r in con.execute(
        "SELECT * FROM oven_records ORDER BY id DESC LIMIT 500").fetchall()]
    con.close()
    return JSONResponse(rows)

@app.get("/oven/api/records/all")
async def oven_api_records_all():
    con = oven_get_db()
    rows = [dict(r) for r in con.execute("SELECT * FROM oven_records ORDER BY id DESC").fetchall()]
    con.close()
    return JSONResponse(rows)

@app.post("/oven/api/record")
async def oven_api_save(request: Request):
    d = await request.json()
    try:
        con = oven_get_db()
        con.execute('''INSERT INTO oven_records
            (date,time,shift,slot,checker,checker_id,confirm_checker,
             machine,machine_code,line,seq,lot,item_fg,item_type,
             set_temp,tol,actual_temp,temp_diff,temp_result,remark,status,saved_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            d.get('date'),       d.get('time'),
            d.get('shift'),      d.get('slot'),
            d.get('checker'),    d.get('checkerID'),    d.get('confirmChecker'),
            d.get('machine'),    d.get('machineCode'),
            d.get('line'),       d.get('seq'),          d.get('lot'),
            d.get('itemFG'),     d.get('type'),
            d.get('setTemp'),    d.get('tol'),          d.get('actualTemp'),
            d.get('tempDiff'),   d.get('tempResult'),
            d.get('remark'),     d.get('status'),
            datetime.now().isoformat()
        ))
        con.commit()
        rid = con.execute('SELECT last_insert_rowid()').fetchone()[0]
        con.close()
        return JSONResponse({'ok': True, 'id': rid})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.put("/oven/api/record/{rid}")
async def oven_api_update(rid: int, request: Request):
    d = await request.json()
    allowed = ['date','time','shift','slot','checker','checker_id','confirm_checker',
               'machine','machine_code','line','seq','lot','item_fg','item_type',
               'set_temp','tol','actual_temp','temp_diff','temp_result','remark','status']
    sets = ', '.join(f'{c}=?' for c in allowed)
    vals = [d.get(c, '') for c in allowed] + [rid]
    try:
        con = oven_get_db()
        con.execute(f'UPDATE oven_records SET {sets} WHERE id=?', vals)
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.delete("/oven/api/record/{rid}")
async def oven_api_delete(rid: int):
    try:
        con = oven_get_db()
        con.execute('DELETE FROM oven_records WHERE id=?', (rid,))
        con.commit(); con.close()
        return JSONResponse({'ok': True})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.post("/oven/api/records/bulk-delete")
async def oven_api_bulk_delete(request: Request):
    try:
        ids = (await request.json()).get('ids', [])
        if not ids: return JSONResponse({'ok': False, 'error': 'No IDs provided'})
        con = oven_get_db()
        placeholders = ','.join('?' * len(ids))
        con.execute(f'DELETE FROM oven_records WHERE id IN ({placeholders})', ids)
        con.commit(); con.close()
        return JSONResponse({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)

@app.get("/oven/api/export")
async def oven_api_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        con = oven_get_db()
        rows = [dict(r) for r in con.execute('SELECT * FROM oven_records ORDER BY id').fetchall()]
        con.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Oven Oil First Lot'
        headers = ['No','Date','Time','Shift','Slot','Checker','Confirm',
                   'Machine','Line','Seq','Lot','Item FG','Type',
                   'Set Temp','Tol ±','Actual','Δ Temp','Temp','Remark','Status']
        ws.append(headers)
        thin = Side(style='thin', color='B0BEC5')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='BF360C')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center')
        pf = PatternFill('solid', fgColor='C8E6C9')
        nf = PatternFill('solid', fgColor='FFCDD2')
        for i, r in enumerate(rows, 1):
            ws.append([i, r.get('date'), r.get('time'),
                'Night' if r.get('shift') == 'night' else 'Day', r.get('slot'),
                r.get('checker'), r.get('confirm_checker'),
                r.get('machine'), r.get('line'), r.get('seq'), r.get('lot'),
                r.get('item_fg'), r.get('item_type'),
                r.get('set_temp'), r.get('tol'), r.get('actual_temp'),
                r.get('temp_diff'), r.get('temp_result'),
                r.get('remark'), r.get('status')])
            for cell in ws[i+1]: cell.border = bdr
            st = str(r.get('status') or '')
            ws.cell(i+1, len(headers)).fill = (
                pf if 'PASS' in st else (nf if 'FAIL' in st else PatternFill()))
        for idx, w in enumerate(
            [5,11,10,8,6,18,18,22,6,6,8,14,8,9,7,8,8,7,20,9], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = ws.dimensions
        out = r"W:\PD\2.HEAT INDOOR\13.Suphamat P\First lot machine\Oven Oil\oven_export.xlsx"
        os.makedirs(os.path.dirname(out), exist_ok=True)
        wb.save(out)
        return JSONResponse({'ok': True, 'file': out, 'rows': len(rows)})
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


# ── Run ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ip = "10.235.117.215"
    print("\n  ================================================")
    print("   First Lot Check Server — port 3001")
    print(f"   Expander  : http://{ip}:3001/expander/firstlot")
    print(f"   Fin Press : http://{ip}:3001/fp/firstlot")
    print(f"   Cutting   : http://{ip}:3001/cutting/firstlot")
    print(f"   HP Check  : http://{ip}:3001/hp/firstlot")
    print(f"   HP Insert : http://{ip}:3001/hp_insert/firstlot")
    print(f"   Oven Oil  : http://{ip}:3001/oven/firstlot")
    print(f"   Dashboard : http://{ip}:3001/dashboard/firstlot")
    print(f"   FP Dash   : http://{ip}:3001/dashboard/firstlot  (shared)")
    print("  ================================================\n")
    uvicorn.run("server_expander:app", host="0.0.0.0", port=3001, reload=False)
